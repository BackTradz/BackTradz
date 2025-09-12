# Script analyseur pour analyser les résultats du runner

 ###==== Analyseur V5

from pathlib import Path
import json
import os
import shutil
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from backend.utils.pip_registry import get_pip  
 

def get_session(hour):
    if 0 <= hour < 8:
        return "Asia"
    elif 8 <= hour < 16:
        return "London"
    else:
        return "New York"



def analyze_file(csv_path, export_dir, STRATEGY_NAME,  symbol, sl_pips, period):
    print("📥 analyse_file lancée")
    print("📂 CSV fourni :", csv_path)
    print("📁 Export dans :", export_dir)
    print("📌 Symbol reçu :", symbol)

    # ✅ Source de vérité: registre central + fallback identique à ton ancien comportement
    s = (symbol or "").upper()
    pip_factor = get_pip(s)
    if pip_factor is None:
        if s.endswith("JPY"):
            pip_factor = 0.01
        elif s == "XAUUSD" or "GC=F" in s or "GOLD" in s:
            pip_factor = 0.1
        elif s.startswith("^"):
            pip_factor = 1.0
        elif s.endswith("-USD") and s.split("-")[0] == "BTC":
            pip_factor = 1.0
        else:
            pip_factor = 0.0001

    print("⚙️ pip_factor :", pip_factor)
    # Lecture CSV
    try:
        df = pd.read_csv(csv_path)
    except pd.errors.EmptyDataError:
        print(f"Fichier vide ignoré: {csv_path}")
        return

    # Skip si pas assez de lignes
    if len(df) < 5:
        print(f"❌ Pas assez de données dans : {csv_path} ({len(df)} lignes)")
        with open("skipped.txt", "a") as log:
            log.write(f"{csv_path} - seulement {len(df)} lignes\n")
        return

    # Préparation des colonnes
    df["time"] = pd.to_datetime(df["time"], errors="coerce")
    df.dropna(subset=["time"], inplace=True)
    df["hour"] = df["time"].dt.hour
    df["date"] = df["time"].dt.date
    df["day_name"] = df["time"].dt.day_name()

    if "sl_size" in df.columns:
        df["sl_size_pips"] = (df["sl_size"] / pip_factor).round(1)
    if "tp1_size" in df.columns:
        df["tp1_size_pips"] = (df["tp1_size"] / pip_factor).round(1)

    df_tp1 = df[df["phase"] == "TP1"]
    df_tp2 = df[df["phase"] == "TP2"]

    total_trades = len(df_tp1)
    tp = (df_tp1["result"] == "TP1").sum()
    sl = (df_tp1["result"] == "SL").sum()
    none = (df_tp1["result"] == "NONE").sum()
    winrate = round((tp / (tp + sl)) * 100, 2) if (tp + sl) > 0 else 0

    buy_count = (df_tp1["direction"] == "buy").sum()
    sell_count = (df_tp1["direction"] == "sell").sum()
    buy_pct = round(buy_count / total_trades * 100, 2)
    sell_pct = round(sell_count / total_trades * 100, 2)

    buy_winrate = round((df_tp1[(df_tp1["direction"] == "buy") & (df_tp1["result"] == "TP1")].shape[0] /
                         df_tp1[df_tp1["direction"] == "buy"].shape[0]) * 100, 2) if buy_count > 0 else 0
    sell_winrate = round((df_tp1[(df_tp1["direction"] == "sell") & (df_tp1["result"] == "TP1")].shape[0] /
                          df_tp1[df_tp1["direction"] == "sell"].shape[0]) * 100, 2) if sell_count > 0 else 0

    avg_sl_size = round(df_tp1["sl_size_pips"].mean(), 1) if "sl_size_pips" in df_tp1 else None
    avg_tp1_size = round(df_tp1["tp1_size_pips"].mean(), 1) if "tp1_size_pips" in df_tp1 else None
    avg_rr_tp1 = round(df_tp1["rr_tp1"].mean(), 2) if "rr_tp1" in df_tp1 else None
    avg_rr_tp2 = round(df_tp2["rr_tp2"].mean(), 2) if "rr_tp2" in df_tp2 else None

    df_tp1 = df_tp1.copy()
    df_tp1["session"] = df_tp1["hour"].apply(get_session)

    # SESSION STATS sécurisé
    session_stats = df_tp1.groupby("session")["result"].value_counts().unstack(fill_value=0)
    for col in ["TP1", "TP2", "SL", "entry"]:
        if col not in session_stats.columns:
            session_stats[col] = 0
    session_stats["total"] = session_stats.sum(axis=1)
    session_stats["winrate"] = (session_stats["TP1"] / session_stats["total"] * 100).round(1)
    session_stats = session_stats[["TP1", "SL", "total", "winrate"]].reset_index()

    # GLOBAL STATS
    global_stats = pd.DataFrame({
        "Metric": [
            "Total Trades", "TP1", "SL", "NONE",
            "Winrate Global", "% Buy", "% Sell",
            "Buy Winrate", "Sell Winrate",
            "SL Size (avg, pips)", "TP1 Size (avg, pips)", "RR TP1 (avg)", "RR TP2 (avg)"
        ],
        "Value": [
            total_trades, tp, sl, none,
            winrate, buy_pct, sell_pct,
            buy_winrate, sell_winrate,
            avg_sl_size, avg_tp1_size, avg_rr_tp1, avg_rr_tp2
        ]
    })

    # HOURLY sécurisé
    hourly = df_tp1.groupby("hour")["result"].value_counts().unstack(fill_value=0)
    for col in ["TP1", "TP2", "SL", "entry"]:
        if col not in hourly.columns:
            hourly[col] = 0
    hourly["total"] = hourly.sum(axis=1)
    hourly["winrate"] = (hourly["TP1"] / hourly["total"] * 100).round(1)
    hourly = hourly[["TP1", "SL", "total", "winrate"]].reset_index()

    # DAY STATS sécurisé
    day_summary = df_tp1.groupby("day_name")["result"].value_counts().unstack(fill_value=0)
    for col in ["TP1", "TP2", "SL", "entry"]:
        if col not in day_summary.columns:
            day_summary[col] = 0
    day_summary["total"] = day_summary.sum(axis=1)
    day_summary["winrate"] = (day_summary["TP1"] / day_summary["total"] * 100).round(1)
    day_summary = day_summary[["TP1", "SL", "total", "winrate"]].reset_index()

    # TP2 STATS
    tp2_total = len(df_tp2)
    tp2_tp = (df_tp2["result"] == "TP2").sum()
    tp2_sl = (df_tp2["result"] == "SL").sum()
    tp2_none = (df_tp2["result"] == "NONE").sum()
    tp2_winrate = round((tp2_tp / total_trades) * 100, 2)  # ✅ basé sur tous les trades


    tp2_stats = pd.DataFrame({
        "Metric": ["Total Trades", "TP2", "SL", "NONE", "Winrate TP2"],
        "Value": [tp2_total, tp2_tp, tp2_sl, tp2_none, tp2_winrate]
    })

    # Nom de fichier dynamique
    xlsx_filename = f"analyse_{STRATEGY_NAME}_{symbol}_SL{sl_pips}_{period}_resultats.xlsx"

    # CSVs intermédiaires (optionnel à renommer aussi si tu veux)
    global_stats.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_global.csv"), index=False)
    session_stats.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_sessions.csv"), index=False)
    hourly.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_par_heure.csv"), index=False)
    day_summary.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_jour_semaine.csv"), index=False)
    tp2_stats.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_tp2_global.csv"), index=False)

    # Excel final unique
    with pd.ExcelWriter(os.path.join(export_dir, xlsx_filename)) as writer:
        global_stats.to_excel(writer, sheet_name="Global", index=False)
        session_stats.to_excel(writer, sheet_name="Sessions", index=False)
        hourly.to_excel(writer, sheet_name="Par_Heure", index=False)
        day_summary.to_excel(writer, sheet_name="Jour_Semaine", index=False)
        tp2_stats.to_excel(writer, sheet_name="TP2_Global", index=False)

    print(f"✅ Analyse terminée pour : {STRATEGY_NAME}")


    # ✅ Ajouter la feuille "Config" avec les paramètres du JSON (robuste)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment

        # 🔎 On prend le params*.json le plus récent (ex: params_0000007.json s'il existe)
        meta_files = sorted(Path(export_dir).glob("params*.json"),
                            key=lambda p: p.stat().st_mtime, reverse=True)
        json_path = meta_files[0] if meta_files else None

        params = {}
        timeframe = "?"
        strategy = STRATEGY_NAME
        pair = symbol
        per = period
        run_seq = ""
        run_id = ""
        user_id = ""

        if json_path:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            params    = data.get("params", {}) or {}
            timeframe = data.get("timeframe", timeframe)
            strategy  = data.get("strategy", strategy)
            pair      = data.get("pair", pair)
            per       = data.get("period", per)
            # infos de run si présentes
            rs = data.get("run_seq")
            run_seq = (str(rs).zfill(7) if rs not in (None, "") else "")
            run_id  = data.get("run_id") or ""
            user_id = data.get("user_id") or ""

        xlsx_path = os.path.join(export_dir, xlsx_filename)
        wb = load_workbook(xlsx_path)

        # (Re)crée la feuille "Config" en première position
        if "Config" in wb.sheetnames:
            ws_old = wb["Config"]
            wb.remove(ws_old)
        ws = wb.create_sheet("Config", 0)

        # En-tête
        ws["A1"], ws["B1"] = "Paramètre", "Valeur"
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center")
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].alignment = center_align
        ws["B1"].alignment = center_align

        # Bloc infos principales
        rows = [
            ("Stratégie", strategy),
            ("Paire",     pair),
            ("Timeframe", timeframe),
            ("Période",   per),
            ("SL / TP",   f"{sl_pips} / {params.get('tp1_pips', '?')} / {params.get('tp2_pips', '?')}"),
            ("pip_factor", str(pip_factor)),
            ("run_seq",   run_seq or "—"),
            ("run_id",    run_id  or "—"),
            ("user_id",   user_id or "—"),
        ]
        r = 2
        for k, v in rows:
            ws[f"A{r}"] = k
            ws[f"B{r}"] = v
            r += 1

        # Ligne vide
        r += 1

        # Détail des paramètres de stratégie
        for k in sorted(params.keys()):
            v = params[k]
            ws[f"A{r}"] = str(k)
            ws[f"B{r}"] = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)
            r += 1

        # Mise en forme colonnes
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 60

        # ✅ Feuille Config stylée ajoutée : (déjà présent)
        wb.save(xlsx_path)
        print("✅ Feuille Config stylée ajoutée :", xlsx_path)
        


        # [MOVE vRUN] Dupliquer l'Excel APRÈS ajout de "Config", pour que la copie l'inclue aussi
        try:
            meta_files = sorted(Path(export_dir).glob("params*.json"),
                                key=lambda p: p.stat().st_mtime, reverse=True)
            json_path = meta_files[0] if meta_files else None
            if json_path:
                with open(json_path, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                run_seq = str(meta.get("run_seq") or "").strip()
                if run_seq:
                    base = os.path.join(export_dir, xlsx_filename)
                    suffixed = os.path.join(
                        export_dir,
                        xlsx_filename.replace("_resultats.xlsx", f"_r{run_seq}_resultats.xlsx")
                    )
                    import shutil as _sh
                    _sh.copyfile(base, suffixed)
        except Exception:
            # silencieux : pas de régression si pas de run_seq
            pass


    except Exception as e:
        print("❌ Impossible d’ajouter la feuille Config :", e)


if __name__ == "__main__":
    analyze_all()
