"""
File: backend/routes/user_dashboard_routes.py
Role: Récupère les backtests d'un utilisateur pour affichage dans son dashboard.
Depends:
  - backend.auth.get_current_user (auth X-API-Key)
  - backend/data/analysis/<...>/*.json (métadonnées de run)
  - openpyxl (lecture 'Global' pour winrate)
Side-effects:
  - Lecture disque (JSON + XLSX)
Security:
  - Protégé par get_current_user (header X-API-Key)
Notes:
  - Les prints '✅/📄/⛔️/📊' sont des logs debug utiles, conservés.
"""

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from pathlib import Path
from app.core.paths import ANALYSIS_DIR, DATA_ROOT
from app.auth import get_current_user, get_user_by_token  # get_user_by_token si besoin
import json
from typing import List
import openpyxl
import os
from fastapi.responses import JSONResponse
import shutil




router = APIRouter()

@router.get("/user/backtests")
def get_user_backtests(request: Request, user=Depends(get_current_user)):
    """
    Scanne backend/data/analysis/** pour les runs de l'utilisateur connecté
    et renvoie des items prêts à afficher (incluant metrics détaillées).
    """
    print(f"✅ USER CONNECTÉ → {user.id}")
    base_dir = ANALYSIS_DIR
    alt_dir = Path("backend/data/analysis").resolve()

    search_dirs = []
    if base_dir.exists():
        search_dirs.append(base_dir)
    if alt_dir.exists() and alt_dir != base_dir:
        search_dirs.append(alt_dir)

    backtests = []
    seen_folders = set()

    # 📂 PROD UNIQUEMENT (Render disk)
    search_root = ANALYSIS_DIR  # ex: /var/data/backtradz/analysis

    # Parcours récursif de params.json
    for meta_path in search_root.glob("**/params.json"):
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))

            folder_name = meta_path.parent.name
            # dédup par nom de dossier (sécurité, si des hardlinks existent)
            if folder_name in seen_folders:
                continue
            seen_folders.add(folder_name)

            print(f"📄 SCANNING {meta_path} | user_id in file: {data.get('user_id')}")
            if data.get("user_id") != user.id:
                continue  # on ne montre que les runs de l'utilisateur courant

            # Champs principaux (on respecte le fichier s'il les fournit)
            symbol    = data.get("pair") or ""
            timeframe = (data.get("timeframe") or "").upper()
            period    = data.get("period") or ""
            strategy  = data.get("strategy") or ""
            sl_pips   = (data.get("params") or {}).get("sl_pips", 100)

            # Fallbacks discrets depuis le nom de dossier si manquants
            try:
                parts = folder_name.split("_")
                if not symbol and len(parts) >= 1:
                    symbol = parts[0]
                if not timeframe and len(parts) >= 2 and parts[1].upper() in {"M1","M5","M15","M30","H1","H4","D1"}:
                    timeframe = parts[1].upper()
                if not period:
                    for p in parts:
                        if "to" in p and any(ch.isdigit() for ch in p):
                            period = p.replace("__", "_")
                            break
            except Exception:
                pass

            # ⛑️ Fallback TF si toujours vide (certaines *upload_custom*)
            if not timeframe:
                timeframe = "H1"

            # Résolution du XLSX (nom standard d'abord, sinon 1er 'analyse_*_resultats.xlsx')
            xlsx_name = f"analyse_{strategy}_{symbol}_SL{sl_pips}_{period}_resultats.xlsx"
            xlsx_path = meta_path.parent / xlsx_name
            if not xlsx_path.exists():
                candidates = list(meta_path.parent.glob("analyse_*_resultats.xlsx"))
                if candidates:
                    xlsx_path = candidates[0]
                    xlsx_name = xlsx_path.name

            # Lecture métriques (tolérante, pas bloquante)
            winrate = "N/A"
            trades  = None
            metrics_payload = None
            if xlsx_path and xlsx_path.exists():
                try:
                    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                    if "Global" in wb.sheetnames:
                        ws = wb["Global"]
                        metrics = {}
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                            key = (str(row[0].value) if row[0].value is not None else "").strip()
                            val = row[1].value if len(row) > 1 else None
                            if key:
                                metrics[key] = val

                        def _get(*keys):
                            for k in keys:
                                if k in metrics and metrics[k] is not None:
                                    return metrics[k]
                            low = {k.lower().replace(" ", ""): k for k in metrics.keys()}
                            for k in keys:
                                t = k.lower().replace(" ", "")
                                if t in low:
                                    return metrics[low[t]]
                            return None

                        def _pct(x):
                            if x is None: return None
                            s = str(x).strip().replace(",", ".")
                            try:
                                n = float(s);  return f"{n*100:.2f}%" if n <= 1 else f"{n:.2f}%"
                            except:
                                return s if s.endswith("%") else f"{s}%"

                        def _num(x):
                            if x is None: return None
                            s = str(x).strip().replace(",", ".")
                            try:
                                return round(float(s), 2)
                            except:
                                return None

                        def _int(x):
                            if x is None: return None
                            s = str(x).strip().replace(",", ".")
                            try:
                                return int(float(s))
                            except:
                                return None

                        wr = _get("Winrate Gl","Winrate Global","WinrateGL","WinrateGlobal","Winrate TP1")
                        if wr is not None: winrate = _pct(wr)
                        tr = _get("Total Trades","Total Trad")
                        if tr is not None: trades = _int(tr)

                        metrics_payload = {
                            "winrate_global": _pct(_get("Winrate Gl","Winrate Global","WinrateGL","WinrateGlobal")),
                            "buy_winrate":    _pct(_get("Buy Winrate","BuyWinrate","Winrate Buy","WinrateBuy")),
                            "sell_winrate":   _pct(_get("Sell Winrate","SellWinrate","Winrate Sell","WinrateSell")),
                            "pct_buy":        _pct(_get("% Buy","%Buy")),
                            "pct_sell":       _pct(_get("% Sell","%Sell")),
                            "tp1":            _int(_get("TP1")),
                            "sl":             _int(_get("SL")),
                            "rr_tp1":         _num(_get("RR TP1 (avg)","RR TP1")),
                            "rr_tp2":         _num(_get("RR TP2 (avg)","RR TP2")),
                            "sl_size":        _num(_get("SL Size (avg,pips)", "SL Size", "Avg SL size", "SL (avg size)", "SL size avg")),
                            "tp1_size":       _num(_get("TP1 Size (avg,pips)", "TP1 Size", "Avg TP1 size", "TP1 (avg size)", "TP1 size avg")),
                        }
                    wb.close()
                except Exception as e:
                    print(f"⚠️ Lecture XLSX échouée ({xlsx_path}) : {e}")

            # Ajout item — structure inchangée (front compatible)
            backtests.append({
                "strategy": strategy,
                "symbol": symbol,
                "timeframe": timeframe,
                "period": period,
                "winrate": winrate,
                "trades": trades,
                "xlsx_filename": xlsx_name,
                "folder": str(meta_path.parent.name),
                "metrics": metrics_payload,
            })

        except Exception as e:
            print(f"❌ Erreur lecture {meta_path} → {e}")
            continue


    print(f"📊 TOTAL BACKTESTS TROUVÉS : {len(backtests)}")
    return backtests




@router.delete("/user/backtests/delete/{folder}")
def delete_backtest(folder: str, user=Depends(get_current_user)):
    """
    Supprime un dossier d'analyse uniquement si l'utilisateur est propriétaire.

    Args:
      folder (str): nom du dossier à supprimer (UUID ou nom généré)

    Returns:
      dict: message de confirmation
    """

    target_path = (ANALYSIS_DIR / folder).resolve()

    if not target_path.exists() or not target_path.is_dir():
        raise HTTPException(status_code=404, detail="Dossier introuvable")

    json_files = list(target_path.glob("*.json"))
    if not json_files:
        raise HTTPException(status_code=400, detail="Aucun fichier JSON trouvé dans le dossier")

    try:
        with open(json_files[0], "r") as f:
            data = json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture JSON: {e}")

    if data.get("user_id") != user.id:
        raise HTTPException(status_code=403, detail="Non autorisé à supprimer ce backtest")

    try:
        shutil.rmtree(target_path)
        print(f"🗑️ Dossier supprimé : {folder}")
        return JSONResponse(content={"message": "Backtest supprimé avec succès"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur suppression dossier: {e}")

# backend/routes/user_dashboard_routes.py


# === CSV deletion endpoints (tolérants) ===============================
from fastapi import Query, Body
from fastapi.responses import Response

def _normalize_backend_rel(p: str) -> str:
    x = (p or "").replace("\\", "/").strip()
    return x


def _delete_csv_file(rel_after_backend: str, user):
    """Supprime le fichier situé sous backend/<rel_after_backend>."""
    # Autorise uniquement des fichiers sous DATA_ROOT (output / output_live / analysis)
    target = Path(rel_after_backend)
    if not target.is_absolute():
        target = (DATA_ROOT / rel_after_backend).resolve()
    if DATA_ROOT not in target.parents and target != DATA_ROOT:
        raise HTTPException(status_code=400, detail="Chemin invalide")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="CSV introuvable")

    try:
        target.unlink()
        print(f"🗑️ CSV supprimé : {target}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur suppression: {e}")

@router.delete("/delete_csv_by_path", status_code=204)
def delete_csv_by_query(path: str = Query(..., description="Chemin incluant 'backend/'"),
                        user=Depends(get_current_user)):
    rel = _normalize_backend_rel(path)
    _delete_csv_file(rel, user)
    return Response(status_code=204)

@router.post("/delete_csv_by_path", status_code=204)
def delete_csv_by_post(payload: dict = Body(...), user=Depends(get_current_user)):
    rel = _normalize_backend_rel(payload.get("path", ""))
    if not rel:
        raise HTTPException(status_code=400, detail="Paramètre 'path' manquant")
    _delete_csv_file(rel, user)
    return Response(status_code=204)

@router.delete("/delete_csv_by_path/{path:path}", status_code=204)
def delete_csv_by_pathparam(path: str, user=Depends(get_current_user)):
    rel = _normalize_backend_rel(path)
    _delete_csv_file(rel, user)
    return Response(status_code=204)


