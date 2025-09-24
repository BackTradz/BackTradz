"""
File: backend/routes/admin_routes.py
Role: Expose des endpoints Admin (statistiques, gestion utilisateurs, historiques).
Depends:
  - backend.models.users.USERS_FILE (JSON de stockage utilisateurs)
  - backend.models.users.get_user_by_token (lookup utilisateur par token)
  - pandas pour l’agrégation de résultats de backtests
Data:
  - Répertoires d’analyses: backend/data/analysis/<symbol>_<tf>_<strategy>_<period...>
Security:
  - Accès réservé à l’admin (contrôle par e-mail dans admin_required()).
  - Auth par header 'X-API-Key' (token utilisateur admin).
Notes:
  - Aucune modification de logique: uniquement docstrings/commentaires.
  - À terme, remplacer les JSON par une base (SQLite/Postgres) et mettre un RBAC propre.
"""


# en haut de fichier (si pas déjà importé)
from fastapi import Body
from app.core.paths import DB_DIR
from app.core.paths import OUTPUT_DIR, ANALYSIS_DIR, USERS_JSON, INVOICES_DIR
from app.core.paths import DATA_ROOT
from fastapi import APIRouter
from fastapi import Request, HTTPException
from app.models.users import USERS_FILE
from fastapi import Request, HTTPException
from pydantic import BaseModel
from app.models.users import get_user_by_token, USERS_FILE
from app.core.admin import require_admin, require_admin_from_request_or_query
from app.models.offers import OFFERS
import os
import pandas as pd
import json
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
import openpyxl
import shutil
from fastapi.responses import FileResponse
from typing import Optional
from fastapi import Body

PARIS_TZ = ZoneInfo("Europe/Paris")

# === AUDIT LEDGER (append-only) ===
AUDIT_DIR = (DATA_ROOT / "audit").resolve()
AUDIT_DIR.mkdir(parents=True, exist_ok=True)
AUDIT_FILE = (AUDIT_DIR / "ledger.jsonl").resolve()

# 📂 Dossier des factures (même que l’émetteur d’invoices)
#  → cf. backend/utils/invoice_generator.py qui écrit dans INVOICES_DIR
FACTURES_DIR = INVOICES_DIR.resolve()
FACTURES_DIR.mkdir(parents=True, exist_ok=True)

def _safe_under_data_root(p: Path) -> bool:
    try:
        return str(p.resolve()).startswith(str(DATA_ROOT.resolve()))
    except Exception:        
        return False

def _factures_stats():
    count = 0
    size  = 0
    for root, _, files in os.walk(FACTURES_DIR):
        for fn in files:
            count += 1
            try:
                size += (Path(root) / fn).stat().st_size
            except Exception:
                pass
    return {"count": count, "bytes": int(size)}


def _infer_subscription_price_from_user(u: dict, tx: dict):
    label  = (tx.get("label") or "").lower()
    method = (tx.get("method") or "").lower()
    reason = (tx.get("billing_reason") or "").lower()
    looks_like_sub = (method in {"renewal", "stripe"} or "subscription" in reason or "abonnement" in label)
    if not looks_like_sub:
        return None
    plan = (u.get("subscription") or {}).get("type") or u.get("plan")
    if not plan:
        return None
    offer = OFFERS.get(plan)
    if not offer or offer.get("type") != "subscription":
        return None
    return float(offer.get("price_eur") or 0)


def _audit_append(obj: dict):
    """N'échoue jamais: on ne bloque pas la requête admin si disque KO."""
    try:
        import json as _json
        from datetime import datetime as _dt, timezone as _tz
        with open(AUDIT_FILE, "a", encoding="utf-8") as f:
            obj = {**obj, "ts": _dt.now(_tz.utc).isoformat().replace("+00:00", "Z")}
            f.write(_json.dumps(obj, ensure_ascii=False) + "\n")
    except Exception:
        pass



router = APIRouter()




@router.get("/admin/ping")
def admin_ping(request: Request):
    require_admin(request)
    return {"ok": True}

@router.get("/admin/factures_info")
def admin_factures_info(request: Request):
    """Infos rapides sur le dossier 'factures'."""
    require_admin_from_request_or_query(request)
    return {"ok": True, **_factures_stats()}

@router.post("/admin/reset-factures")
def admin_reset_factures(request: Request):
    """
    Vide le dossier 'factures' sur le disk (suppression récursive des fichiers).
    Sécurité: ne supprime que sous DATA_ROOT/factures.
    """
    require_admin_from_request_or_query(request)
    if not _safe_under_data_root(FACTURES_DIR):
        raise HTTPException(status_code=400, detail="Chemin non autorisé.")

    deleted_files = 0
    deleted_dirs  = 0
    try:
        for child in FACTURES_DIR.iterdir():
            try:
                if child.is_file():
                    child.unlink(missing_ok=True)
                    deleted_files += 1
                elif child.is_dir():
                    shutil.rmtree(child, ignore_errors=True)
                    deleted_dirs += 1
            except Exception:
                # on continue, on ne casse pas l'endpoint
                pass
    except Exception:
        raise HTTPException(status_code=500, detail="Échec du nettoyage.")

    stats = _factures_stats()
    return {"ok": True, "deleted_files": deleted_files, "deleted_dirs": deleted_dirs, "left": stats}

@router.get("/admin/factures_list")
def admin_factures_list(request: Request, limit: int = 200):
    """Liste des fichiers du dossier factures (nom, taille, mtime)."""
    require_admin_from_request_or_query(request)
    items = []
    try:
        for p in FACTURES_DIR.glob("**/*"):
            if p.is_file():
                st = p.stat()
                items.append({
                    "name": p.name,
                    "rel": str(p.relative_to(FACTURES_DIR)),
                    "bytes": int(st.st_size),
                    "mtime": datetime.fromtimestamp(st.st_mtime).isoformat(),
                })
        # tri recent -> ancien
        items.sort(key=lambda x: x["mtime"], reverse=True)
    except Exception:
        items = []
    return {"ok": True, "items": items[:max(1, min(limit, 1000))]}


@router.get("/admin/factures_download")
def admin_factures_download(request: Request, rel: str):
    """Télécharge un fichier de facture depuis le disk (admin-only)."""
    # accepte Authorization en header OU ?apiKey=... en query (comme /admin/download_xlsx)
    require_admin_from_request_or_query(request)
    target = (FACTURES_DIR / rel).resolve()
    if not _safe_under_data_root(target) or not str(target).startswith(str(FACTURES_DIR.resolve())):
        raise HTTPException(status_code=400, detail="Chemin non autorisé.")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable.")
    # Forcer le download
    return FileResponse(
        path=target,
        media_type="application/octet-stream",
        filename=target.name,
        headers={"Content-Disposition": f'attachment; filename="{target.name}"'}
    )


@router.post("/admin/factures_delete")
def admin_factures_delete(request: Request, rel: Optional[str] = Body(None), payload: dict = Body(None)):
    """Supprime un fichier unique dans le dossier factures (admin-only)."""
    require_admin_from_request_or_query(request)
    # tolérant: body JSON {"rel": "..."} OU query ?rel=...
    if rel is None:
        rel = (payload or {}).get("rel") if payload else None
    if rel is None:
        rel = request.query_params.get("rel")
    if not rel:
        raise HTTPException(status_code=400, detail="Paramètre 'rel' manquant.")
    target = (FACTURES_DIR / rel).resolve()
    if not _safe_under_data_root(target) or not str(target).startswith(str(FACTURES_DIR.resolve())):
        raise HTTPException(status_code=400, detail="Chemin non autorisé.")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable.")
    try:
        target.unlink(missing_ok=False)
    except Exception:
        raise HTTPException(status_code=500, detail="Échec de suppression.")
    # renvoyer stats à jour pour rafraîchir l'UI
    return {"ok": True, "left": _factures_stats()}


@router.get("/admin/factures_delete")
def admin_factures_delete_get(request: Request, rel: str):
    """
    Variante GET pour supprimer une facture (même contrat que download):
    /api/admin/factures_delete?rel=...&apiKey=...
    Retourne aussi les stats restantes pour faciliter le refresh UI.
    """
    require_admin_from_request_or_query(request)
    target = (FACTURES_DIR / rel).resolve()
    if not _safe_under_data_root(target) or not str(target).startswith(str(FACTURES_DIR.resolve())):
        raise HTTPException(status_code=400, detail="Chemin non autorisé.")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable.")
    try:
        target.unlink(missing_ok=False)
    except Exception:
        raise HTTPException(status_code=500, detail="Échec de suppression.")
    # ✅ aligne le contrat sur la POST: renvoyer les stats restantes
    return {"ok": True, "left": _factures_stats()}


@router.get("/admin/stats/daily_summary")
def daily_summary():
    """
    Agrège, par jour, TP1/TP2/SL à partir des CSV de backtests.

    Parcourt chaque dossier dans 'backend/data/analysis', détecte:
    - symbol, timeframe (tf), strategy depuis le nom du dossier,
    - un CSV 'backtest_result.csv',
    - un JSON '...settings/result...' (facultatif) pour récupérer 'params' et 'time_key'.

    Retour:
        list[dict]: [{ date, symbol, timeframe, strategy, tp1, tp2, sl, params }, ...]

    Notes:
        - Détection auto de la colonne temps si 'time_key' absent
          (essaie 'entry_time' → 'Datetime' → 'time').
        - Si les colonnes tp1_hit/tp2_hit/sl_hit n’existent pas, on les calcule
          à partir des colonnes 'phase' et 'result'.
        - Ignore silencieusement les dossiers mal nommés / CSV invalides (logs print).
    """
    base_path = ANALYSIS_DIR
    summary = []

    for folder in base_path.iterdir():
        if not folder.is_dir():
            continue

        folder_name = folder.name
        try:
            parts = folder_name.split("_")
            symbol, tf, strategy = parts[0], parts[1], "_".join(parts[2:-2])
        except Exception:
            # Dossier non conforme au pattern attendu → on ignore
            print(f"⚠️ Dossier ignoré (nom incorrect) : {folder_name}")
            continue

        csv_path = folder / "backtest_result.csv"
        if not csv_path.exists():
            continue

        # 🔍 Cherche le fichier JSON (params + time_key éventuels)
        json_path = None
        for file in folder.glob("*.json"):
            if "settings" in file.name or "result" in file.name:
                json_path = file
                break

        params = {}
        time_key = None

        # 📄 Lecture params + time_key depuis le JSON (si dispo)
        if json_path and json_path.exists():
            try:
                with open(json_path, "r") as f:
                    json_data = json.load(f)
                    params = json_data.get("params", {}) or {}
                    time_key = params.get("time_key")
            except Exception as e:
                print(f"⚠️ JSON invalide : {e}")

        # 🔍 Auto-fallback si time_key non fourni ou colonne manquante
        try:
            df_sample = pd.read_csv(csv_path, nrows=1)
            if not time_key or time_key not in df_sample.columns:
                for candidate in ["entry_time", "Datetime", "time"]:
                    if candidate in df_sample.columns:
                        time_key = candidate
                        break
        except Exception as e:
            print(f"❌ Erreur lecture partielle CSV : {e}")
            continue

        if not time_key:
            print(f"❌ Aucune colonne de temps trouvée dans {csv_path.name}")
            continue

        # ✅ Lecture complète + conversion date
        try:
            df = pd.read_csv(csv_path)
            df[time_key] = pd.to_datetime(df[time_key], errors="coerce")
            df = df.dropna(subset=[time_key])
            df["date"] = df[time_key].dt.date

            # ✅ Création des flags TP1 / TP2 / SL si manquants
            if not {"tp1_hit", "tp2_hit", "sl_hit"}.issubset(df.columns):
                # NB: On déduit les flags à partir des champs phase/result
                df["tp1_hit"] = (df.get("phase") == "TP1") & (df.get("result") == "TP1")
                df["tp2_hit"] = (df.get("phase") == "TP2") & (df.get("result") == "TP2")
                df["sl_hit"]  = (df.get("result") == "SL")
        except Exception as e:
            print(f"❌ Erreur lecture CSV complet {csv_path.name}: {e}")
            continue

        # ✅ Groupement par jour (somme booléens → compte)
        daily = df.groupby("date").agg({
            "tp1_hit": "sum",
            "tp2_hit": "sum",
            "sl_hit": "sum"
        }).reset_index()

        for _, row in daily.iterrows():
            summary.append({
                "date": str(row["date"]),
                "symbol": symbol,
                "timeframe": tf,
                "strategy": strategy,
                "tp1": int(row["tp1_hit"]),
                "tp2": int(row["tp2_hit"]),
                "sl": int(row["sl_hit"]),
                "params": params
            })

    return summary

@router.get("/admin/stats/backtest_summary")
def backtest_summary():
    """
    Résumé global par dossier de backtest (total/TP1/TP2/SL + winrates).

    Pour chaque dossier 'backend/data/analysis/...':
      - Déduit symbol/tf/strategy/period depuis le nom,
      - Lit 'backtest_result.csv' (+ JSON params/time_key si présent),
      - Calcule total/TP1/TP2/SL, et winrate TP1/TP2,
      - Essaie de lire les tailles depuis l'XLSX (feuille Global) si dispo :
        SL Size (avg, pips), TP1 Size (avg, pips), TP2 Size (avg, pips)
        (ou TP Size (avg, pips) comme fallback global).
    """
    base_path = ANALYSIS_DIR
    summary = []

    for folder in base_path.iterdir():
        if not folder.is_dir():
            continue

        folder_name = folder.name
        # ✂️ si le dossier finit par "__h<hash>", on retire proprement ce suffixe
        base_name = folder_name.split("__h", 1)[0]
        try:
            parts = base_name.split("_")
            symbol, tf = parts[0], parts[1]
            strategy = "_".join(parts[2:-2]) if len(parts) >= 5 else "_".join(parts[2:-1])

            # 🧠 Reconstruction robuste de la période:
            #  1) préférer le segment qui contient "to" (ex: "2025-07-01to2025-07-31")
            period = ""
            for p in parts:
                if "to" in p and any(ch.isdigit() for ch in p):
                    period = p.replace("__", "_")
                    break
            #  2) fallback legacy si non trouvée
            if not period and len(parts) >= 4:
                last = parts[-1]
                prev = parts[-2]
                period = f"{prev}to{last.replace('sl','').replace('.zip','')}"
        except Exception:
            print(f"❌ Dossier ignoré (mauvais nom) : {folder_name}")
            continue

        csv_path = folder / "backtest_result.csv"
        json_path = next(folder.glob("*.json"), None)
        if not csv_path.exists():
            continue

        # --- params + time_key depuis JSON (optionnel) ---
        params, time_key = {}, None
        backtest_id = None
        if json_path and json_path.exists():
            try:
                with open(json_path, "r") as f:
                    json_data = json.load(f)
                    params = json_data.get("params", {}) or {}
                    time_key = params.get("time_key")
                    # id tolérant (selon ce que tes JSON contiennent)
                    backtest_id = (
                        json_data.get("id")
                        or json_data.get("backtest_id")
                        or json_data.get("run_id")
                    )
            except Exception:
                pass

        # --- auto-détection time_key si manquant ---
        try:
            df_sample = pd.read_csv(csv_path, nrows=1)
            if not time_key or time_key not in df_sample.columns:
                for candidate in ["entry_time", "Datetime", "time"]:
                    if candidate in df_sample.columns:
                        time_key = candidate
                        break
        except Exception:
            continue
        if not time_key:
            print(f"⛔ Aucune colonne de temps détectée dans : {folder_name}")
            continue

        # --- lecture CSV + flags ---
        try:
            df = pd.read_csv(csv_path)
            df[time_key] = pd.to_datetime(df[time_key], errors="coerce")
            df = df.dropna(subset=[time_key])

            if not {"tp1_hit", "tp2_hit", "sl_hit"}.issubset(df.columns):
                df["tp1_hit"] = (df.get("phase") == "TP1") & (df.get("result") == "TP1")
                df["tp2_hit"] = (df.get("phase") == "TP2") & (df.get("result") == "TP2")
                df["sl_hit"]  = (df.get("result") == "SL")
        except Exception as e:
            print(f"⛔ Erreur lecture CSV complet : {folder_name} → {e}")
            continue

        # --- agrégations de base (cohérent avec le dashboard) ---
        df_tp1 = df[df["phase"] == "TP1"] if "phase" in df.columns else df
        total = int(len(df_tp1))
        tp1 = int(len(df_tp1[df_tp1["result"] == "TP1"])) if "result" in df_tp1.columns else 0
        tp2 = int(len(df[(df.get("phase") == "TP2") & (df.get("result") == "TP2")])) if "phase" in df.columns and "result" in df.columns else 0
        sl  = int(len(df_tp1[df_tp1["result"] == "SL"])) if "result" in df_tp1.columns else 0

        winrate_tp1 = round(tp1 / total * 100, 1) if total > 0 else 0.0
        winrate_tp2 = round(tp2 / total * 100, 1) if total > 0 else 0.0
        tp1_rate = round(tp1 / total * 100, 1) if total > 0 else 0.0
        tp2_rate = round(tp2 / total * 100, 1) if total > 0 else 0.0
        sl_rate  = round(sl  / total * 100, 1) if total > 0 else 0.0

        rr_tp1 = rr_tp2 = None
        if "rr_tp1" in df.columns:
            try: rr_tp1 = round(float(df["rr_tp1"].mean()), 2)
            except: pass
        if "rr_tp2" in df.columns:
            try: rr_tp2 = round(float(df["rr_tp2"].mean()), 2)
            except: pass

        # ================= LECTURE XLSX =================
        sl_size = tp1_size = tp2_size = None
        try:
            sl_pips = int(params.get("sl_pips") or params.get("sl") or 100)
            xlsx_name = f"analyse_{strategy}_{symbol}_SL{sl_pips}_{period}_resultats.xlsx"
            xlsx_path = folder / xlsx_name

            if not xlsx_path.exists():
                variants = {period}
                if period:
                    variants.update({
                        period.replace("_to_", " to "),
                        period.replace(" to ", "_to_"),
                        period.replace(" ", "_"),
                        period.replace("_", " "),
                    })
                for per in variants:
                    cand = folder / f"analyse_{strategy}_{symbol}_SL{sl_pips}_{per}_resultats.xlsx"
                    if cand.exists():
                        xlsx_path = cand
                        break

            if xlsx_path.exists():
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                if "Global" in wb.sheetnames:
                    ws = wb["Global"]
                    metrics_map = {}
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        key = (str(row[0].value) if row[0].value else "").strip()
                        val = row[1].value if len(row) > 1 else None
                        if key:
                            metrics_map[key] = val

                    # --- valeurs directes ---
                    sl_size  = metrics_map.get("SL Size (avg, pips)")  or metrics_map.get("SL Size (avg,pips)")
                    tp1_size = metrics_map.get("TP1 Size (avg, pips)") or metrics_map.get("TP1 Size (avg,pips)") \
                               or metrics_map.get("TP Size (avg, pips)") or metrics_map.get("TP Size (avg,pips)")

                    # FIX TP2: fallback sur TP Size (global), pas sur TP1
                    tp2_size = metrics_map.get("TP2 Size (avg, pips)") or metrics_map.get("TP2 Size (avg,pips)") \
                               or metrics_map.get("TP Size (avg, pips)") or metrics_map.get("TP Size (avg,pips)")
                wb.close()
        except Exception as e:
            print(f"⚠️ Impossible de lire XLSX pour {folder.name}: {e}")

        # ====== LECTURE XLSX (variante) + normalisation nombres ======
        try:
            sl_pips = int(params.get("sl_pips") or params.get("sl") or 100)
            xlsx_name = f"analyse_{strategy}_{symbol}_SL{sl_pips}_{period}_resultats.xlsx"
            xlsx_path = folder / xlsx_name

            if not xlsx_path.exists():
                variants = {period}
                if period:
                    variants.update({
                        period.replace("_to_", " to "),
                        period.replace(" to ", "_to_"),
                        period.replace(" ", "_"),
                        period.replace("_", " "),
                    })
                for per in variants:
                    cand = folder / f"analyse_{strategy}_{symbol}_SL{sl_pips}_{per}_resultats.xlsx"
                    if cand.exists():
                        xlsx_path = cand
                        break

            if xlsx_path.exists():
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                if "Global" in wb.sheetnames:
                    ws = wb["Global"]
                    metrics_map = {}
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        key = (str(row[0].value) if row[0].value is not None else "").strip()
                        val = row[1].value if len(row) > 1 else None
                        if key:
                            metrics_map[key] = val

                    sl_xlsx       = metrics_map.get("SL Size (avg, pips)")  or metrics_map.get("SL Size (avg,pips)")
                    tp1_xlsx      = metrics_map.get("TP1 Size (avg, pips)") or metrics_map.get("TP1 Size (avg,pips)")
                    tp2_xlsx      = metrics_map.get("TP2 Size (avg, pips)") or metrics_map.get("TP2 Size (avg,pips)")
                    tp_global_xls = metrics_map.get("TP Size (avg, pips)")  or metrics_map.get("TP Size (avg,pips)")

                    def _num(x):
                        if x is None: return None
                        s = str(x).replace(",", ".")
                        try: return float(s)
                        except: return x

                    if sl_xlsx is not None:  sl_size  = _num(sl_xlsx)
                    if tp1_xlsx is not None: tp1_size = _num(tp1_xlsx)
                    if tp2_xlsx is not None: tp2_size = _num(tp2_xlsx)

                    # FIX TP2: fallback sur TP Size (global) si TP2 absent
                    if tp2_size is None and tp_global_xls is not None:
                        tp2_size = _num(tp_global_xls)
                    # idem pour TP1 (si tu veux garder la logique globale)
                    if tp1_size is None and tp_global_xls is not None:
                        tp1_size = _num(tp_global_xls)

                wb.close()
        except Exception as e:
            print(f"⚠️ Impossible de lire XLSX pour {folder.name}: {e}")

        # ====== FALLBACKS éventuels depuis le CSV (si colonnes présentes) ======
        for col in ["SL Size (avg, pips)", "SL Size (avg,pips)", "sl_pips", "sl_points"]:
            if sl_size is None and col in df.columns:
                try: sl_size = round(float(pd.to_numeric(df[col], errors="coerce").mean()), 2)
                except: pass
        for col in ["TP1 Size (avg, pips)", "TP1 Size (avg,pips)", "tp1_pips", "tp1_points"]:
            if tp1_size is None and col in df.columns:
                try: tp1_size = round(float(pd.to_numeric(df[col], errors="coerce").mean()), 2)
                except: pass
        for col in ["TP2 Size (avg, pips)", "TP2 Size (avg,pips)", "tp2_pips", "tp2_points"]:
            if tp2_size is None and col in df.columns:
                try: tp2_size = round(float(pd.to_numeric(df[col], errors="coerce").mean()), 2)
                except: pass
        # TP Size global → fallback pour TP1 et/ou TP2 si encore manquants
        if tp1_size is None or tp2_size is None:
            for col in ["TP Size (avg, pips)", "TP Size (avg,pips)", "tp_size", "tp_pips", "tp_points"]:
                if col in df.columns:
                    try:
                        avg_tp = round(float(pd.to_numeric(df[col], errors="coerce").mean()), 2)
                        if tp1_size is None: tp1_size = avg_tp
                        if tp2_size is None: tp2_size = avg_tp
                        break
                    except:
                        pass

        # ====== DERNIER PASS XLSX générique (aucun clone TP1→TP2) ======
        try:
            xlsx_file = next(folder.glob("analyse_*_resultats.xlsx"), None)
            if xlsx_file:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                if "Global" in wb.sheetnames:
                    ws = wb["Global"]
                    kv = {}
                    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        k = (str(r[0].value) if r[0].value is not None else "").strip()
                        v = r[1].value if len(r) > 1 else None
                        if k:
                            kv[k] = v

                    if sl_size  is None: sl_size  = kv.get("SL Size (avg, pips)")  or kv.get("SL Size (avg,pips)")
                    if tp1_size is None: tp1_size = kv.get("TP1 Size (avg, pips)") or kv.get("TP1 Size (avg,pips)")
                    if tp2_size is None: tp2_size = kv.get("TP2 Size (avg, pips)") or kv.get("TP2 Size (avg,pips)")

                    # FIX TP2: fallback global si toujours manquant
                    if (tp1_size is None or tp2_size is None):
                        tp_global = kv.get("TP Size (avg, pips)") or kv.get("TP Size (avg,pips)")
                        if tp_global is not None:
                            if tp1_size is None: tp1_size = tp_global
                            if tp2_size is None: tp2_size = tp_global
                wb.close()
        except Exception as e:
            print(f"⚠️ XLSX read error in {folder.name}: {e}")

        # --- Push ---
        summary.append({
           "folder": folder.name,      # 👈 identifiant dossier pour overlay & masquage front
            "id": backtest_id,          # 👈 id depuis le JSON si présent (sinon null)
            "symbol": symbol,
            "timeframe": tf,
            "strategy": strategy,
            "period": period,
            "total": total,
            "tp1": tp1,
            "tp2": tp2,
            "sl": sl,
            "winrate_tp1": winrate_tp1,
            "winrate_tp2": winrate_tp2,
            "tp1_rate": tp1_rate,
            "tp2_rate": tp2_rate,
            "sl_rate": sl_rate,
            "rr_tp1": rr_tp1,
            "rr_tp2": rr_tp2,
            "tp1_size": tp1_size,
            "tp2_size": tp2_size,  # ✅ plus de clone TP1
            "sl_size": sl_size,
            "params": params
        })

    return summary

from datetime import datetime, timezone

def _parse_iso(dt):
    try:
        return datetime.fromisoformat(dt.replace("Z", "+00:00"))
    except Exception:
        return None

def _is_failed_payment_tx_local(tx: dict) -> bool:
    lbl = str(tx.get("label") or "").lower()
    status = str(tx.get("status") or "").lower()
    reason = str(tx.get("billing_reason") or "").lower()
    fcode = str(tx.get("failure_code") or "").lower()
    return ("échou" in lbl or "echou" in lbl or "failed" in lbl
            or status in {"failed","payment_failed","failed_payment"}
            or "payment_failed" in reason or bool(fcode))

def _is_subscription_tx_local(u: dict | None, tx: dict) -> bool:
    lbl = str(tx.get("label") or "").lower()
    method = str(tx.get("method") or "").lower()
    reason = str(tx.get("billing_reason") or "").lower()
    if "abonnement" in lbl or "subscription" in reason or method == "renewal":
        return True
    if u:
        plan = (u.get("subscription") or {}).get("type") or u.get("plan") or ""
        return str(plan).upper().startswith("SUB")
    return False

@router.get("/admin/get_users")
def get_all_users(request: Request):
    require_admin(request)
    raw = json.loads(USERS_FILE.read_text(encoding="utf-8"))
    out = []
    now = datetime.utcnow().replace(tzinfo=None)

    for uid, u in raw.items():
        sub = u.get("subscription") or {}
        active_sub = bool(sub.get("active"))
        renew_dt = _parse_iso(sub.get("renew_date") or "")
        last_status = str(sub.get("last_payment_status") or "").lower()

        # Jours de retard (seulement si ABO ACTIF et date dépassée)
        days_overdue = 0
        if active_sub and renew_dt and now > renew_dt.replace(tzinfo=None):
            days_overdue = (now - renew_dt.replace(tzinfo=None)).days

        # Échecs liés à l’ABO (on ignore les one-shot pour l’état "RETARD")
        fails_sub = []
        for tx in (u.get("purchase_history") or []):
            if _is_failed_payment_tx_local(tx) and _is_subscription_tx_local(u, tx):
                fails_sub.append(tx.get("date") or "")
        fails_sub.sort(reverse=True)
        last_failed = fails_sub[0] if fails_sub else ""

        # ⚠️ payment_issue = uniquement si ABO ACTIF et (jours de retard > 0 ou statut failed ou échec abo)
        payment_issue = active_sub and ((days_overdue > 0) or (last_status == "failed") or bool(fails_sub))

        out.append({
            "id": uid,
            "email": u.get("email"),
            "username": u.get("username"),
            "plan": u.get("plan"),
            "credits": u.get("credits"),
            "is_blocked": u.get("is_blocked", False),
            "subscription": sub,
            # indicateurs affichés en table
            "payment_issue": payment_issue,
            "days_overdue": days_overdue,      # ⏱ ex : 4
            "failed_payments": len(fails_sub), # échecs ABO (pas one-shot)
            "last_failed": last_failed,
        })
    return out


class UserAction(BaseModel):
    """
    Payload générique pour les actions admin sur un utilisateur.

    Attribs:
        user_id (str): identifiant (clé) de l'utilisateur dans USERS_FILE.
        amount (int): valeur numérique (ex: +1 crédit); défaut = 1.
    """
    user_id: str
    amount: int = 1  # par défaut pour les crédits

def load_users():
    """Lecture JSON brute de USERS_FILE."""
    with open(USERS_FILE, "r") as f:
        return json.load(f)

def save_users(data):
    """Écriture JSON formatée (indent=2) vers USERS_FILE."""
    with open(USERS_FILE, "w") as f:
        json.dump(data, f, indent=2)

# Remplace l’ancienne implémentation par un simple proxy vers la nouvelle
def admin_required(request: Request):
    return require_admin(request)

@router.post("/admin/add_credit")
def add_credit(payload: UserAction, request: Request):
    """
    Ajoute des crédits à un utilisateur.

    Args:
        payload (UserAction): user_id cible, amount à ajouter.
        request (Request): pour check admin_required().

    Raises:
        404: si l'utilisateur n'existe pas.

    Returns:
        { "detail": "..."} message simple de confirmation.
    """
    admin_required(request)
    users = load_users()
    if payload.user_id not in users:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    users[payload.user_id]["credits"] = users[payload.user_id].get("credits", 0) + payload.amount
    save_users(users)
    return {"detail": f"{payload.amount} crédit(s) ajouté(s)."}

@router.post("/admin/remove_credit")
def remove_credit(payload: UserAction, request: Request):
    """
    Retire des crédits à un utilisateur (plancher 0).

    Args:
        payload (UserAction): user_id cible, amount à retirer.
    """
    admin_required(request)
    users = load_users()
    if payload.user_id not in users:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    users[payload.user_id]["credits"] = max(0, users[payload.user_id].get("credits", 0) - payload.amount)
    save_users(users)
    return {"detail": f"{payload.amount} crédit(s) retiré(s)."}

@router.post("/admin/toggle_block_user")
def toggle_block_user(payload: UserAction, request: Request):
    """
    Bascule l’état 'bloqué' d’un utilisateur (on/off).

    Args:
        payload (UserAction): user_id cible.
    """
    admin_required(request)
    users = load_users()
    if payload.user_id not in users:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    current = users[payload.user_id].get("is_blocked", False)
    users[payload.user_id]["is_blocked"] = not current
    save_users(users)
    return {"detail": f"Utilisateur {'bloqué' if not current else 'débloqué'}."}

    
@router.post("/admin/delete_user")
def delete_user(payload: UserAction, request: Request):
    """
    Supprime définitivement un utilisateur du JSON.

    DANGER:
      - Opération destructive sans corbeille/backup.

    Args:
        payload (UserAction): user_id cible.
    """
    admin_required(request)
    users = load_users()
    uid = payload.user_id
    if uid not in users:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    # 1) Archiver ses transactions (pour stats immuables)
    for tx in (users[uid].get("purchase_history") or []):
        _audit_append({"type": "tx", "user_id": uid, "data": tx})

    # 2) Événement de suppression
    _audit_append({"type": "user_deleted", "user_id": uid})

    # 3) Suppression effective
    del users[uid]
    save_users(users)
    return {"detail": "Utilisateur supprimé définitivement."}
    

@router.get("/admin/user_history/{user_id}")
def get_user_history(user_id: str, request: Request):
    """
    Historique d’achats d’un utilisateur (tri desc sur date), avec montant (€) :
    - price_eur si présent
    - sinon price_paid si > 0
    - sinon prix d’offre (abo) inféré depuis OFFERS
    """
    admin_required(request)

    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            users = json.load(f)
    except:
        raise HTTPException(status_code=500, detail="Erreur lecture users.json")

    u = users.get(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    out = []
    for tx in (u.get("purchase_history") or []):
        # calcul du montant € affichable
        price_eur = tx.get("price_eur")
        if price_eur is None:
            p = tx.get("price_paid")
            if isinstance(p, (int, float)) and p > 0:
                price_eur = float(p)
        if not isinstance(price_eur, (int, float)) or price_eur <= 0:
            price_eur = _infer_subscription_price_from_user(u, tx) or 0.0

        item = dict(tx)  # on conserve tous les champs d’origine
        item["price_eur"] = float(price_eur)
        out.append(item)

    # Tri décroissant par date (support 'date' manquante)
    return sorted(out, key=lambda x: x.get("date", ""), reverse=True)

@router.get("/admin/global_history")
def get_global_history(request: Request):
    """
    Historique global (toutes transactions des utilisateurs).
    Chaque item = { username, date, price_paid, price_eur, method, label, type, credits_delta }
    avec price_eur enrichi si besoin.
    """
    admin_required(request)

    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            users = json.load(f)
    except:
        raise HTTPException(status_code=500, detail="Erreur lecture users.json")

    history = []
    for uid, u in users.items():
        pseudo = u.get("username") or uid
        for tx in (u.get("purchase_history") or []):
            # montant €
            price_eur = tx.get("price_eur")
            if price_eur is None:
                p = tx.get("price_paid")
                if isinstance(p, (int, float)) and p > 0:
                    price_eur = float(p)
            if not isinstance(price_eur, (int, float)) or price_eur <= 0:
                price_eur = _infer_subscription_price_from_user(u, tx) or 0.0

            history.append({
                "username": pseudo,
                "date": tx.get("date"),
                "price_paid": tx.get("price_paid", 0),
                "price_eur": float(price_eur),
                "method": tx.get("method", "inconnu"),
                "label": tx.get("label") or tx.get("offer_id", "Crédits"),
                "type": tx.get("type"),
                "credits_delta": tx.get("credits_delta"),
            })

    return sorted(history, key=lambda x: x.get("date", ""), reverse=True)

@router.get("/admin/global_stats")
def get_global_stats(request: Request):
    """
    Statistiques globales simples à partir de users.json.

    Calcule:
      - total_sales_eur: somme de 'price_paid' sur toutes les transactions.
      - total_credits_bought: compteur d’achats de crédits (hors 'mensuel'/'offert').
      - credits_offered: compteur d’items labelisés "mensuel"/"offert".
      - total_users: nb d’utilisateurs.
      - subscribers: nb d’utilisateurs dont plan commence par 'SUB'.
      - credits_available: somme des crédits restants (users[*].credits).
    """
    admin_required(request)

    try:
        with open(USERS_FILE, "r") as f:
            users = json.load(f)
    except:
        raise HTTPException(status_code=500, detail="Erreur lecture users.json")

    stats = {
        "total_sales_eur": 0,
        "total_credits_bought": 0,
        "credits_offered": 0,
        "total_users": 0,
        "subscribers": 0,
        "credits_available": 0
    }

    for uid, u in users.items():
        stats["total_users"] += 1
        stats["credits_available"] += u.get("credits", 0)

        if u.get("plan", "").startswith("SUB"):
            stats["subscribers"] += 1

        for tx in u.get("purchase_history", []):
            prix = tx.get("price_paid", 0)
            stats["total_sales_eur"] += prix

            label = tx.get("label", "").lower()
            if "mensuel" in label or "offert" in label:
                stats["credits_offered"] += 1
            else:
                stats["total_credits_bought"] += 1

    return stats


from fastapi.responses import FileResponse


@router.get("/admin/download_xlsx")
def admin_download_xlsx(
    request: Request,
    symbol: str,
    timeframe: str,
    strategy: str,
    period: str,
    sl: int = 100,
):
    """
    Télécharge le fichier XLSX d'un backtest (feuille Global).
    Auth: admin (header X-API-Key ou query ?apiKey=...)
    """
    # sécurité admin (source of truth, tolère header X-API-Key et, si activé, ?apiKey=)
    require_admin_from_request_or_query(request)


    base_dir = ANALYSIS_DIR
    # --- Normalisation de 'period' : garde only 'YYYY-MM-DDtoYYYY-MM-DD'
    #    (certains dossiers/period traînent 'to100' / '_100' / '_sl100')
    import re
    m = re.search(r"(20\d{2}-\d{2}-\d{2})\s*to\s*(20\d{2}-\d{2}-\d{2})", period.replace("_", " "))
    clean_period = f"{m.group(1)}to{m.group(2)}" if m else period

    # 1) Sélectionne les dossiers plausibles (symbol + timeframe + strategy)
    candidates = []
    for folder in base_dir.iterdir():
        if not folder.is_dir():
            continue
        name = folder.name
        if not name.startswith(f"{symbol}_{timeframe}_"):
            continue
        if strategy and strategy not in name:
            continue
        candidates.append(folder)
    if not candidates:
        raise HTTPException(status_code=404, detail="Dossier introuvable pour ces paramètres")

    # 2) Résout le nom du XLSX (tolérant aux espaces/_ dans period)
    def resolve_xlsx(folder: Path):
        filename = f"analyse_{strategy}_{symbol}_SL{sl}_{clean_period}_resultats.xlsx"
        path = folder / filename
        if path.exists():
            return path

        variants = {
            clean_period,
            clean_period.replace("_to_", " to "),
            clean_period.replace(" to ", "_to_"),
            clean_period.replace(" ", "_"),
            clean_period.replace("_", " "),
        }
        for per in variants:
            cand = folder / f"analyse_{strategy}_{symbol}_SL{sl}_{per}_resultats.xlsx"
            if cand.exists():
                return cand

        # fallback permissif
        cands = list(folder.glob(f"analyse_{strategy}_{symbol}_SL*_*resultats.xlsx")) \
                or list(folder.glob("analyse_*_resultats.xlsx"))
        return cands[0] if cands else None

    for folder in candidates:
        xlsx = resolve_xlsx(folder)
        if xlsx and xlsx.exists():
            return FileResponse(
                path=str(xlsx),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=xlsx.name,
            )

    raise HTTPException(status_code=404, detail="Fichier XLSX introuvable")

from fastapi import Body

@router.post("/admin/history/reset")
def reset_history(request: Request, scope: str = Body("all")):
    """
    Reset de l'historique (ADMIN).
    scope: "all" | "sales" | "backtests"
      - all       : vide complètement purchase_history
      - sales     : supprime uniquement les achats/renouvellements (laisse les backtests)
      - backtests : supprime uniquement les lignes de backtest
    ⚠️ Ne touche pas aux abonnements, crédits ou autres champs.
    """
    admin_required(request)
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            users = json.load(f)
    except:
        raise HTTPException(status_code=500, detail="Erreur lecture users.json")

    changed = 0
    for uid, u in users.items():
        ph = u.get("purchase_history") or []
        if scope == "all":
            if ph:
                u["purchase_history"] = []
                changed += 1
        elif scope == "sales":
            keep = []
            for tx in ph:
                ttype = (tx.get("type") or "").lower()
                label = (tx.get("label") or "").lower()
                is_backtest = (
                    ttype == "backtest"
                    or "backtest" in label
                    or (tx.get("symbol") and tx.get("timeframe") and tx.get("strategy"))
                )
                if is_backtest:
                    keep.append(tx)
            if len(keep) != len(ph):
                u["purchase_history"] = keep
                changed += 1
        elif scope == "backtests":
            keep = []
            for tx in ph:
                ttype = (tx.get("type") or "").lower()
                label = (tx.get("label") or "").lower()
                is_backtest = (
                    ttype == "backtest"
                    or "backtest" in label
                    or (tx.get("symbol") and tx.get("timeframe") and tx.get("strategy"))
                )
                if not is_backtest:
                    keep.append(tx)
            if len(keep) != len(ph):
                u["purchase_history"] = keep
                changed += 1

    if changed:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, ensure_ascii=False, indent=2)

    return {"status": "ok", "changed_users": changed}


#---- on reutilise la meme route--- 
RECREATE_FILE = Path(USERS_FILE).parent / "email_recreate.json"

def _load_json_safe(path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

@router.get("/admin/email-recreate")
def admin_get_email_recreate(request: Request):
    """Retourne le contenu actuel du compteur (clé email -> nb tentatives)."""
    require_admin(request)
    if not RECREATE_FILE.exists():
        return {"ok": True, "counts": {}}
    return {"ok": True, "counts": _load_json_safe(RECREATE_FILE)}

@router.post("/admin/reset-email-recreate")
def admin_reset_email_recreate(request: Request, payload: dict | None = Body(None)):
    """
    Reset du compteur de créations d'email.
    - payload={"email":"x@y.com"} -> supprime cette entrée uniquement
    - payload=None/{} -> reset total du fichier
    """
    require_admin(request)

    if not RECREATE_FILE.exists():
        return {"ok": True, "message": "Fichier inexistant (déjà vide)."}

    counts = _load_json_safe(RECREATE_FILE)
    email = (payload or {}).get("email")
    if email:
        email_norm = email.strip().lower()
        if email_norm in counts:
            counts.pop(email_norm, None)
            RECREATE_FILE.write_text(json.dumps(counts, indent=2), encoding="utf-8")
            return {"ok": True, "cleared": email_norm, "left": counts}
        return {"ok": True, "message": f"Aucune entrée pour {email_norm}."}

    # reset total
    RECREATE_FILE.write_text("{}", encoding="utf-8")
    return {"ok": True, "cleared": "ALL"}
