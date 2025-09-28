# backend/routes/backtest_xlsx_routes.py
"""
Rôle
----
Expose des routes READ-ONLY pour lire un fichier .xlsx d'un backtest:
- /user/backtests/xlsx/meta       → métadonnées (feuilles, dimensions, colonnes)
- /user/backtests/xlsx/sheet      → lecture paginée d'une feuille (offset/limit)
- /user/backtests/xlsx/aggregates → agrégats (overall, par heure, par session, par jour)

Sécurité
--------
- Protégé par get_current_user (X-API-Key).
- Vérifie que le backtest (dossier) appartient à l'utilisateur (lecture JSON).
- Lecture disque en "data/analysis/…".

Perf / Robustesse
-----------------
- Lecture "data_only=True" (formules évaluées).
- Pagination pour /sheet (offset / limit) → évite de charger 200k lignes.
- Heuristiques tolérantes (noms de colonnes courants).
- Cache simple en mémoire (optionnel) : non inclus ici pour rester minimal et sûr.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse
from pathlib import Path
from typing import Optional, List, Dict, Any
from app.auth import get_current_user
import openpyxl
import json
import os
from datetime import datetime
import math
from app.services.backtest_xlsx_service import (
    _is_admin, _analysis_base, _folder_path, _assert_owns_folder,
    _guess_xlsx_path, _safe_str, _to_dt, _session_for_hour
)
from app.core.admin import is_admin_user  # ✅ source of truth admin


router = APIRouter()

# -------- Routes ----------------------------------------------------------

@router.get("/user/backtests/xlsx/meta")
def xlsx_meta(
    folder: str = Query(..., description="Nom du dossier d'analyse (ex: AUDUSD_M5_...)"),
    user=Depends(get_current_user)
):
    """
    Renvoie les métadonnées du classeur:
    {
      "filename": "...xlsx",
      "sheets": [{ "name": "Global", "rows": 42, "cols": 8, "columns": ["A","B",...] }, ...]
    }
    """
    folder_dir = _assert_owns_folder(folder, user)
    xlsx_path = _guess_xlsx_path(folder_dir)
    if not xlsx_path or not xlsx_path.exists():
        raise HTTPException(404, "Fichier .xlsx introuvable")

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            # tentative de colonnes depuis la première ligne (si c'est un header)
            columns = []
            if rows >= 1:
                first = [c for c in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
                columns = [(_safe_str(x).strip() or f"col{i+1}") for i, x in enumerate(first)]
            sheets.append({
                "name": name,
                "rows": rows,
                "cols": cols,
                "columns": columns,
            })
        wb.close()
        return {
            "filename": xlsx_path.name,
            "sheets": sheets,
        }
    except Exception as e:
        raise HTTPException(500, f"Erreur ouverture .xlsx: {e}")

@router.get("/user/backtests/xlsx/sheet")
def xlsx_sheet(
    folder: str = Query(...),
    sheet: str = Query(..., description="Nom exact de la feuille (ex: Global, Trades, ...)"),
    offset: int = Query(0, ge=0),
    limit: int = Query(500, ge=1, le=5000),
    use_header: int = Query(1, description="1: map par colonnes de la 1ère ligne, 0: renvoie des arrays"),
    user=Depends(get_current_user),
):
    """
    Lecture paginée d'une feuille.
    - use_header=1: renvoie des objets {col:val} sur la base de la 1ère ligne comme en-tête
    - use_header=0: renvoie des listes (valeurs brutes)
    """
    folder_dir = _assert_owns_folder(folder, user)
    xlsx_path = _guess_xlsx_path(folder_dir)
    if not xlsx_path or not xlsx_path.exists():
        raise HTTPException(404, "Fichier .xlsx introuvable")

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(404, f"Feuille '{sheet}' introuvable")

        ws = wb[sheet]
        total = ws.max_row or 0
        start_row = 2 if use_header else 1  # si header, on saute la 1ère ligne
        # borne la fenêtre
        first = start_row + offset
        last = min(total, first + limit - 1)
        if first > total:
            rows = []
        else:
            it = ws.iter_rows(min_row=first, max_row=last, values_only=True)
            rows = [list(r) for r in it]

        # header éventuel
        columns = []
        if use_header and total >= 1:
            header = [c for c in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
            columns = [(_safe_str(x).strip() or f"col{i+1}") for i, x in enumerate(header)]
            # map en dict
            mapped = []
            for r in rows:
                obj = {}
                for i, v in enumerate(r):
                    key = columns[i] if i < len(columns) else f"col{i+1}"
                    obj[key] = v
                mapped.append(obj)
            rows = mapped

        wb.close()

        has_more = (last < total)
        next_offset = offset + limit if has_more else None

        return {
            "total_rows": total - (1 if use_header and total > 0 else 0),
            "offset": offset,
            "limit": limit,
            "has_more": has_more,
            "next_offset": next_offset,
            "columns": columns,
            "rows": rows
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Erreur lecture feuille: {e}")

# ---- Agrégats (overall, par heure, par session, par jour) ---------------

def _to_dt(v) -> Optional[datetime]:
    # openpyxl renvoie soit datetime, soit serial excel/texte; on gère minimalement
    if isinstance(v, datetime):
        return v
    s = _safe_str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S%z", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None

def _session_for_hour(h: int) -> str:
    # sessions simplifiées (UTC): ajuste si besoin
    # Asia ~ 0–7, London ~ 7–13, NY ~ 13–21, Late ~ 21–24
    if   0 <= h < 7:  return "Asia"
    elif 7 <= h < 13: return "London"
    elif 13 <= h < 21:return "NY"
    else:             return "Late"

@router.get("/user/backtests/xlsx/aggregates")
def xlsx_aggregates(
    folder: str = Query(...),
    sheet: str = Query("Trades", description="Feuille contenant les opérations (ex: Trades)"),
    dt_col: str = Query("Datetime"),
    r_col: str  = Query("R", description="Colonne du résultat en R (>0 = win)"),
    group: str  = Query("overall", regex="^(overall|hour|session|weekday)$"),
    user=Depends(get_current_user),
):
    """
    Calcule des agrégats simples depuis la feuille "Trades" (ou autre):
    - overall : winrate, expectancyR (moyenne de R), profit factor, trades
    - hour    : winrate par heure
    - session : winrate par session (Asia/London/NY/Late)
    - weekday : winrate par jour (Lun..Dim)
    """
    folder_dir = _assert_owns_folder(folder, user)
    xlsx_path = _guess_xlsx_path(folder_dir)
    if not xlsx_path or not xlsx_path.exists():
        raise HTTPException(404, "Fichier .xlsx introuvable")

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(404, f"Feuille '{sheet}' introuvable")

        ws = wb[sheet]
        if ws.max_row < 2:
            wb.close()
            return {"overall": {"trades": 0}}

        # header
        header = [c for c in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        hidx = { (_safe_str(k).strip() or f"col{i+1}"): i for i, k in enumerate(header) }

        if dt_col not in hidx or r_col not in hidx:
            wb.close()
            raise HTTPException(400, f"Colonnes requises manquantes: {dt_col}, {r_col}")

        i_dt = hidx[dt_col]; i_r = hidx[r_col]

        # accumulateurs
        trades = 0
        wins = 0
        losses = 0
        sum_pos = 0.0
        sum_neg = 0.0
        sum_r   = 0.0

        by_hour: Dict[int, Dict[str, Any]] = {}
        by_sess: Dict[str, Dict[str, Any]] = {}
        by_wd  : Dict[int, Dict[str, Any]] = {}

        # lecture lignes
        it = ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)
        for row in it:
            dt = _to_dt(row[i_dt])
            try:
                r  = float(str(row[i_r]).replace(",", "."))
            except Exception:
                continue
            trades += 1
            sum_r += r
            if r > 0: wins += 1; sum_pos += r
            elif r < 0: losses += 1; sum_neg += abs(r)

            if group in ("hour","session","weekday") and dt:
                h = dt.hour
                sess = _session_for_hour(h)
                wd = (dt.weekday() + 1)  # 1=Lundi..7=Dimanche

                if group == "hour":
                    b = by_hour.setdefault(h, {"trades":0,"wins":0,"sumR":0.0})
                    b["trades"] += 1; b["sumR"] += r; 
                    if r>0: b["wins"] += 1

                if group == "session":
                    b = by_sess.setdefault(sess, {"trades":0,"wins":0,"sumR":0.0})
                    b["trades"] += 1; b["sumR"] += r; 
                    if r>0: b["wins"] += 1

                if group == "weekday":
                    b = by_wd.setdefault(wd, {"trades":0,"wins":0,"sumR":0.0})
                    b["trades"] += 1; b["sumR"] += r; 
                    if r>0: b["wins"] += 1

        wb.close()

        # overall
        wr = (wins / trades) if trades else 0.0
        expectancy = (sum_r / trades) if trades else 0.0
        pf = (sum_pos / sum_neg) if sum_neg > 0 else (math.inf if sum_pos>0 else 0.0)

        if group == "overall":
            return {
                "overall": {
                    "trades": trades,
                    "winrate": wr,
                    "expectancyR": round(expectancy, 4),
                    "pf": None if math.isinf(pf) else round(pf, 4),
                }
            }

        if group == "hour":
            out = []
            for h in sorted(by_hour.keys()):
                b = by_hour[h]; t=b["trades"]; w=b["wins"]
                out.append({
                    "hour": h, "trades": t,
                    "winrate": (w/t) if t else 0.0,
                    "avgR": round((b["sumR"]/t), 4) if t else 0.0
                })
            return {"by_hour": out}

        if group == "session":
            order = ["Asia","London","NY","Late"]
            out = []
            for s in order:
                if s in by_sess:
                    b = by_sess[s]; t=b["trades"]; w=b["wins"]
                    out.append({
                        "session": s, "trades": t,
                        "winrate": (w/t) if t else 0.0,
                        "avgR": round((b["sumR"]/t), 4) if t else 0.0
                    })
            return {"by_session": out}

        if group == "weekday":
            labels = {1:"Lundi",2:"Mardi",3:"Mercredi",4:"Jeudi",5:"Vendredi",6:"Samedi",7:"Dimanche"}
            out = []
            for d in range(1,8):
                if d in by_wd:
                    b = by_wd[d]; t=b["trades"]; w=b["wins"]
                    out.append({
                        "weekday": d, "label": labels[d],
                        "trades": t,
                        "winrate": (w/t) if t else 0.0,
                        "avgR": round((b["sumR"]/t), 4) if t else 0.0
                    })
            return {"by_weekday": out}

        # fallback
        return {"overall": {"trades": trades, "winrate": wr}}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Erreur calcul agrégats: {e}")
