# Script analyseur pour analyser les résultats du runner

 ###==== Analyseur V5

from pathlib import Path
import json
import os
import shutil
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns


 

def get_session(hour):
    if 0 <= hour < 8:
        return "Asia"
    elif 8 <= hour < 16:
        return "London"
    else:
        return "New York"

def get_pip_value(symbol):
    symbol = symbol.upper().replace("/", "").replace("-", "")
    if "BTC" in symbol or "XBT" in symbol:
        return 0.01
    elif "ETH" in symbol:
        return 0.01
    elif "JPY" in symbol:
        return 0.01
    elif "XAU" in symbol or "GC=F" in symbol or "GOLD" in symbol:
        return 0.1
    return 0.0001



def analyze_file(csv_path, export_dir, STRATEGY_NAME,  symbol, sl_pips, period):
    print("📥 analyse_file lancée")
    print("📂 CSV fourni :", csv_path)
    print("📁 Export dans :", export_dir)
    print("📌 Symbol reçu :", symbol)

    pip_factor = get_pip_value(symbol)

    print("⚙️ pip_factor :", pip_factor)
    # Lecture CSV
    try:
        df = pd.read_csv(csv_path)
    except pd.errors.EmptyDataError:
        print(f"Fichier vide ignoré: {csv_path}")
        return

    # Skip si pas assez de lignes
    if len(df) < 5:
        print(f"❌ Pas assez de données dans : {csv_path} ({len(df)} lignes)")
        with open("skipped.txt", "a") as log:
            log.write(f"{csv_path} - seulement {len(df)} lignes\n")
        return

    # Préparation des colonnes
    df["time"] = pd.to_datetime(df["time"], errors="coerce")
    df.dropna(subset=["time"], inplace=True)
    df["hour"] = df["time"].dt.hour
    df["date"] = df["time"].dt.date
    df["day_name"] = df["time"].dt.day_name()

    if "sl_size" in df.columns:
        df["sl_size_pips"] = (df["sl_size"] / pip_factor).round(1)
    if "tp1_size" in df.columns:
        df["tp1_size_pips"] = (df["tp1_size"] / pip_factor).round(1)

    df_tp1 = df[df["phase"] == "TP1"]
    df_tp2 = df[df["phase"] == "TP2"]

    total_trades = len(df_tp1)
    tp = (df_tp1["result"] == "TP1").sum()
    sl = (df_tp1["result"] == "SL").sum()
    none = (df_tp1["result"] == "NONE").sum()
    winrate = round((tp / (tp + sl)) * 100, 2) if (tp + sl) > 0 else 0

    buy_count = (df_tp1["direction"] == "buy").sum()
    sell_count = (df_tp1["direction"] == "sell").sum()
    buy_pct = round(buy_count / total_trades * 100, 2)
    sell_pct = round(sell_count / total_trades * 100, 2)

    buy_winrate = round((df_tp1[(df_tp1["direction"] == "buy") & (df_tp1["result"] == "TP1")].shape[0] /
                         df_tp1[df_tp1["direction"] == "buy"].shape[0]) * 100, 2) if buy_count > 0 else 0
    sell_winrate = round((df_tp1[(df_tp1["direction"] == "sell") & (df_tp1["result"] == "TP1")].shape[0] /
                          df_tp1[df_tp1["direction"] == "sell"].shape[0]) * 100, 2) if sell_count > 0 else 0

    avg_sl_size = round(df_tp1["sl_size_pips"].mean(), 1) if "sl_size_pips" in df_tp1 else None
    avg_tp1_size = round(df_tp1["tp1_size_pips"].mean(), 1) if "tp1_size_pips" in df_tp1 else None
    avg_rr_tp1 = round(df_tp1["rr_tp1"].mean(), 2) if "rr_tp1" in df_tp1 else None
    avg_rr_tp2 = round(df_tp2["rr_tp2"].mean(), 2) if "rr_tp2" in df_tp2 else None

    df_tp1 = df_tp1.copy()
    df_tp1["session"] = df_tp1["hour"].apply(get_session)

    # SESSION STATS sécurisé
    session_stats = df_tp1.groupby("session")["result"].value_counts().unstack(fill_value=0)
    for col in ["TP1", "TP2", "SL", "entry"]:
        if col not in session_stats.columns:
            session_stats[col] = 0
    session_stats["total"] = session_stats.sum(axis=1)
    session_stats["winrate"] = (session_stats["TP1"] / session_stats["total"] * 100).round(1)
    session_stats = session_stats[["TP1", "SL", "total", "winrate"]].reset_index()

    # GLOBAL STATS
    global_stats = pd.DataFrame({
        "Metric": [
            "Total Trades", "TP1", "SL", "NONE",
            "Winrate Global", "% Buy", "% Sell",
            "Buy Winrate", "Sell Winrate",
            "SL Size (avg, pips)", "TP1 Size (avg, pips)", "RR TP1 (avg)", "RR TP2 (avg)"
        ],
        "Value": [
            total_trades, tp, sl, none,
            winrate, buy_pct, sell_pct,
            buy_winrate, sell_winrate,
            avg_sl_size, avg_tp1_size, avg_rr_tp1, avg_rr_tp2
        ]
    })

    # HOURLY sécurisé
    hourly = df_tp1.groupby("hour")["result"].value_counts().unstack(fill_value=0)
    for col in ["TP1", "TP2", "SL", "entry"]:
        if col not in hourly.columns:
            hourly[col] = 0
    hourly["total"] = hourly.sum(axis=1)
    hourly["winrate"] = (hourly["TP1"] / hourly["total"] * 100).round(1)
    hourly = hourly[["TP1", "SL", "total", "winrate"]].reset_index()

    # DAY STATS sécurisé
    day_summary = df_tp1.groupby("day_name")["result"].value_counts().unstack(fill_value=0)
    for col in ["TP1", "TP2", "SL", "entry"]:
        if col not in day_summary.columns:
            day_summary[col] = 0
    day_summary["total"] = day_summary.sum(axis=1)
    day_summary["winrate"] = (day_summary["TP1"] / day_summary["total"] * 100).round(1)
    day_summary = day_summary[["TP1", "SL", "total", "winrate"]].reset_index()

    # TP2 STATS
    tp2_total = len(df_tp2)
    tp2_tp = (df_tp2["result"] == "TP2").sum()
    tp2_sl = (df_tp2["result"] == "SL").sum()
    tp2_none = (df_tp2["result"] == "NONE").sum()
    tp2_winrate = round((tp2_tp / total_trades) * 100, 2)  # ✅ basé sur tous les trades


    tp2_stats = pd.DataFrame({
        "Metric": ["Total Trades", "TP2", "SL", "NONE", "Winrate TP2"],
        "Value": [tp2_total, tp2_tp, tp2_sl, tp2_none, tp2_winrate]
    })

    # Nom de fichier dynamique
    xlsx_filename = f"analyse_{STRATEGY_NAME}_{symbol}_SL{sl_pips}_{period}_resultats.xlsx"

    # CSVs intermédiaires (optionnel à renommer aussi si tu veux)
    global_stats.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_global.csv"), index=False)
    session_stats.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_sessions.csv"), index=False)
    hourly.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_par_heure.csv"), index=False)
    day_summary.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_jour_semaine.csv"), index=False)
    tp2_stats.to_csv(os.path.join(export_dir, f"{STRATEGY_NAME}_tp2_global.csv"), index=False)

    # Excel final unique
    with pd.ExcelWriter(os.path.join(export_dir, xlsx_filename)) as writer:
        global_stats.to_excel(writer, sheet_name="Global", index=False)
        session_stats.to_excel(writer, sheet_name="Sessions", index=False)
        hourly.to_excel(writer, sheet_name="Par_Heure", index=False)
        day_summary.to_excel(writer, sheet_name="Jour_Semaine", index=False)
        tp2_stats.to_excel(writer, sheet_name="TP2_Global", index=False)

    print(f"✅ Analyse terminée pour : {STRATEGY_NAME}")

     # ✅ Ajouter la feuille "Config" avec les paramètres du JSON
    try:
        from openpyxl import load_workbook

        json_path = next(Path(export_dir).glob("*.json"))
        with open(json_path, "r") as f:
            data = json.load(f)
            params = data.get("params", {})
            timeframe = data.get("timeframe", "?")
            strategy = data.get("strategy", STRATEGY_NAME)
            pair = data.get("pair", symbol)
            period = data.get("period", period)

        xlsx_path = os.path.join(export_dir, xlsx_filename)
        from openpyxl.styles import Font, Alignment

        # Chargement
        wb = load_workbook(xlsx_path)
        ws = wb.create_sheet("Config", 0)

        # En-tête
        ws["A1"] = "Paramètre"
        ws["B1"] = "Valeur"
        header_font = Font(bold=True)   
        center_align = Alignment(horizontal="center")

        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].alignment = center_align
        ws["B1"].alignment = center_align

        # Données principales
        rows = [
            ("Stratégie", strategy),
            ("Paire", pair),
            ("Timeframe", timeframe),
            ("Période", period),
            ("SL / TP", f"{params.get('sl_pips', '?')} / {params.get('tp1_pips', '?')} / {params.get('tp2', '?')}")
        ]

        for i, (key, value) in enumerate(rows, start=2):
            ws[f"A{i}"] = key
            ws[f"B{i}"] = value

        # Params dynamiques (avec espace visuel après les 5 premières lignes)
        row_idx = len(rows) + 3

        for k, v in params.items():
            if k not in ["sl_pips", "tp1_pips", "tp2"]:
                ws[f"A{row_idx}"] = k
                ws[f"B{row_idx}"] = str(v)
                row_idx += 1

        # Formatage général
        for col in ["A", "B"]:
            ws.column_dimensions[col].width = 25

        wb.save(xlsx_path)
        print("✅ Feuille Config stylée ajoutée :", xlsx_path)

    
    except Exception as e:
        print("❌ Impossible d’ajouter la feuille Config :", e)

if __name__ == "__main__":
    analyze_all()
