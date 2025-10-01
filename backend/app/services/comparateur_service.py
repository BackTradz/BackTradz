# app/services/comparateur_service.py
"""
Service comparateur:
- Liste d’analyses 'options' pour l'user courant (select multi côté front)
- Construction de séries normalisées à partir des CSV déjà générés:
    *_global.csv
    *_sessions.csv
    *_par_heure.csv
    *_jour_semaine.csv
- Sécurité: filtre par ownership via params*.json (user_id/run_user.id) si présent
"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import json
import pandas as pd
import openpyxl  # fallback XLSX comme le dashboard

from app.core.paths import ANALYSIS_DIR
from app.schemas.comparateur import (
    CompareOptionsItem, CompareOptionsResponse,
    CompareDataRequest, CompareDataResponse, SeriesItem
)
from app.schemas.communs import DEFAULT_SESSIONS, DEFAULT_DAYS, DEFAULT_HOURS
import unicodedata
import re

# ------------ Helpers lecture disque ------------

def _find_runs(root: Path) -> List[Path]:
    """
    Retourne les dossiers d’analyse (runs) sous ANALYSIS_DIR.
    Tolérant: détecte via plusieurs artefacts possibles (prod/dev),
    pour coller au dashboard.
    """
    if not root.exists():
        return []
    candidates: List[Path] = []
    # params (nom exact ou variantes)
    for pat in ("params.json", "params*.json"):
        candidates += [p.parent for p in root.rglob(pat)]
    # xlsx résultats (si pas de params)
    for pat in ("*resultats.xlsx", "analyse_*_resultats.xlsx"):
        candidates += [p.parent for p in root.rglob(pat)]
    # csv globaux (fallback ultime)
    for pat in ("*_global.csv", "global.csv", "Global.csv"):
        candidates += [p.parent for p in root.rglob(pat)]
    runs = sorted(set(candidates), key=lambda d: d.stat().st_mtime, reverse=True)
    return runs


def _load_params(run_dir: Path) -> dict:
    """Lit le params*.json le plus récent s’il existe (meta du run)."""
    meta_files = sorted(run_dir.glob("params*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not meta_files:
        return {}
    try:
        return json.loads(meta_files[0].read_text(encoding="utf-8"))
    except Exception:
        return {}

def _safe_read_csv(path: Path) -> Optional[pd.DataFrame]:
    try:
        return pd.read_csv(path)
    except Exception:
        return None

def _detect_files(run_dir: Path) -> Dict[str, Optional[Path]]:
    # ⚠️ 0 régression : on garde le 'global' historique (TP1),
    #    et on ajoute un slot optionnel 'tp2_global' si présent.
    files = {"global": None, "tp2_global": None, "sessions": None, "hour": None, "day": None}
    # on ignore explicitement les fichiers *_tp2_global.csv pour 'global'
    for f in run_dir.glob("*_global.csv"):
        if not str(f.name).lower().endswith("tp2_global.csv"):
            files["global"] = f
            break
    # détection optionnelle CSV TP2
    for f in run_dir.glob("*_tp2_global.csv"):
        files["tp2_global"] = f
        break
    for f in run_dir.glob("*_sessions.csv"):     files["sessions"] = f; break
    for f in run_dir.glob("*_par_heure.csv"):    files["hour"] = f;     break
    for f in run_dir.glob("*_jour_semaine.csv"): files["day"] = f;      break
    return files

def _own_by_user(params: dict, current_user_id: str) -> bool:
    """
    Aligné dashboard : si user_id présent, on compare de façon permissive.
    - accepte run_user.id ou user_id
    - cast en str, trim, insensible casse
    - si pas d'info user -> on accepte (legacy)
    """
    try:
        raw = (
            (params.get("run_user") or {}).get("id")
            or (params.get("user") or {}).get("id")   # ✅ couvre 'user': {id: ...}
            or params.get("user_id")
            or ""
        )
        uid = str(raw).strip().lower()
        cur = str(current_user_id).strip().lower()
        return (uid == "") or (uid == cur)
    except Exception:
        return True
    
def _clean_symbol(x: str) -> str:
    """Considère UNKNOWN/?/- comme vide pour pair/symbol (aligné dashboard)."""
    x = (x or "").strip()
    return "" if x.upper() in {"UNKNOWN", "?", "-"} else x

def _normalize_pair_symbol(params: dict, run_dir: Path) -> Tuple[str, str]:
    """
    Normalise pair/symbol pour label UI et séries :
    - nettoie UNKNOWN/?/-
    - synchronise pair <-> symbol
    - fallback sur le préfixe du nom de dossier (ex: GC=F_M15_...)
    """
    pair = _clean_symbol(params.get("pair"))
    symbol = _clean_symbol(params.get("symbol"))
    if not pair and symbol:
        pair = symbol
    if not symbol and pair:
        symbol = pair
    if not pair or not symbol:
        folder_code = run_dir.name.split("_")[0].strip()
        if folder_code:
            if not pair:   pair = folder_code
            if not symbol: symbol = folder_code
    return pair, symbol


def _compose_label(params: dict, fallback_dir: Path) -> Tuple[str, str]:
    """Construit (label, period) lisibles pour l’UI."""
    pair = str(params.get("pair") or params.get("symbol") or "")
    period = str(params.get("period") or "").strip() or ""
    tf = str(params.get("timeframe") or "").strip()
    strategy = str(params.get("strategy") or "").strip()
    core = f"{pair} · {period}" if period else pair
    if tf:        core = f"{core} · {tf}"
    if strategy:  core = f"{core} · {strategy}"
    if core.strip() in {"·", "", "?"}:
        core = fallback_dir.name
    return core, (period or "")

def _extract_global_metrics(df_global: pd.DataFrame) -> Tuple[Optional[int], Optional[float], Optional[float]]:
    """
    global.csv a normalement lignes 'Metric' + 'Value'.
    On récupère Total Trades, Winrate Global (TP1) en %.
    """
    try:
        df = df_global
        if not {"Metric", "Value"}.issubset(df.columns):
            return None, None, None
        # Map normalisé -> valeur numérique
        kv = {}
        for _, r in df.iterrows():
            k = _norm_key(r.get("Metric", ""))
            v = r.get("Value", None)
            if v is None:
                continue
            try:
                v = float(str(v).replace("%", "").strip())
            except Exception:
                continue
            if k:
                kv[k] = v

        def pick(*candidates: str) -> Optional[float]:
            # essaie plusieurs clés candidates (normalisées)
            for c in candidates:
                v = kv.get(_norm_key(c))
                if v is not None:
                    return v
            return None

        total_trades   = pick("Total Trades", "Trades", "Nombre Trades", "Total")  # tolérant
        winrate_global = pick("Winrate Global", "Winrate TP1", "TP1 Winrate", "Winrate")  # TP1
        winrate_tp2    = pick("TP2 Winrate", "Winrate TP2")
        trades_count = int(total_trades) if total_trades is not None else None
        winrate_tp1 = (winrate_global / 100.0) if winrate_global is not None else None
        winrate_tp2 = (winrate_tp2 / 100.0) if winrate_tp2 is not None else None
        return trades_count, winrate_tp1, winrate_tp2
    except Exception:
        return None, None, None
    
# ---------- Helper : extraction période depuis un nom de dossier ----------
_PERIOD_RE = re.compile(r"(?P<start>\d{4}-\d{2}-\d{2})\s*to\s*(?P<end>\d{4}-\d{2}-\d{2})", re.IGNORECASE)

def _extract_period_from_dirname(dirname: str) -> str | None:
    """
    Exemples acceptés :
      ..._2025-08-01to2025-08-31_...
      ..._2025-08-01_to_2025-08-31_...
      ... 2025-08-01 to 2025-08-31 ...
    Retourne '01-08-2025 → 31-08-2025' (format EU) ou None.
    """
    m = _PERIOD_RE.search(dirname.replace("_to_", "to").replace("–", "-").replace("—", "-"))
    if not m:
        return None
    start = m.group("start")
    end = m.group("end")
    try:
        from datetime import datetime
        s_dt = datetime.strptime(start, "%Y-%m-%d")
        e_dt = datetime.strptime(end, "%Y-%m-%d")
        start_fmt = s_dt.strftime("%d-%m-%Y")
        end_fmt = e_dt.strftime("%d-%m-%Y")
        return f"{start_fmt} → {end_fmt}"
    except Exception:
        # fallback brut si parsing échoue
        return f"{start} → {end}"
   
# ---------- Helpers de normalisation ----------
def _norm_key(s: str) -> str:
    """
    Normalise un libellé de métrique:
    - retire accents/espaces/underscore/%
    - minuscule
    Ex: "TP2 Winrate" / "winrate TP2" / "Winrate_TP2" => "tp2winrate"
    """
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace(" ", "").replace("_", "").replace("%", "")
    return s.lower()

_DAY_ALIASES = {
    # EN 3 lettres -> liste d'alias acceptés
    "Mon": {"mon", "monday", "lun", "lundi"},
    "Tue": {"tue", "tuesday", "mar", "mardi"},
    "Wed": {"wed", "wednesday", "mer", "mercredi"},
    "Thu": {"thu", "thursday", "jeu", "jeudi"},
    "Fri": {"fri", "friday", "ven", "vendredi"},
    "Sat": {"sat", "saturday", "sam", "samedi"},
    "Sun": {"sun", "sunday", "dim", "dimanche"},
}

def _normalize_day_label(s: str) -> str:
    """
    Retourne le code 3 lettres EN attendu par DEFAULT_DAYS
    à partir de variantes EN/FR.
    """
    key = _norm_key(s)
    for code, variants in _DAY_ALIASES.items():
        if key in variants:
            return code
    # Si déjà au bon format:
    if s in DEFAULT_DAYS:
        return s
    # Dernier recours: garde tel quel (ne cassera pas l'ordre des buckets)
    return s

# ------------ API: options ------------

def list_user_compare_options(current_user_id: str) -> CompareOptionsResponse:
    items: List[CompareOptionsItem] = []
    #print(f"🔎 [compare] ANALYSIS_DIR = {ANALYSIS_DIR}")  # debug non bloquant (même style que dashboard)
    runs = _find_runs(ANALYSIS_DIR)
    #print(f"📦 [compare] dossiers détectés = {len(runs)}")
    for run_dir in runs:
        params = _load_params(run_dir)
        # pair/symbol normalisés (même logique partout)
        pair, symbol = _normalize_pair_symbol(params, run_dir)
         # ownership identique dashboard
        if not _own_by_user(params, current_user_id):
            continue
        
        files = _detect_files(run_dir)
        trades_count, wr1, wr2 = (None, None, None)
        # 1) CSV global si présent
        if files["global"]:
            df_global = _safe_read_csv(files["global"])
            if df_global is not None:
                trades_count, wr1, wr2 = _extract_global_metrics(df_global)

        # 1-bis) CSV TP2 global (optionnel) pour remplir wr2 s'il manque
        if wr2 is None and files.get("tp2_global"):
            df_tp2 = _safe_read_csv(files["tp2_global"])
            if df_tp2 is not None and {"Metric", "Value"}.issubset(df_tp2.columns):
                try:
                    row = df_tp2.loc[df_tp2["Metric"].astype(str) == "Winrate TP2", "Value"]
                    if not row.empty:
                        v = str(row.iloc[0]).replace("%", "").strip()
                        wr2 = float(v)/100.0
                except Exception:
                    pass

        # 2) Fallback XLSX (onglet "Global") si pas de CSV
        if trades_count is None and wr1 is None and wr2 is None:
            try:
                # tentative: 1er fichier 'analyse_*_resultats.xlsx' dans le dossier
                candidates = list(run_dir.glob("analyse_*_resultats.xlsx"))
                if candidates:
                    wb = openpyxl.load_workbook(candidates[0], data_only=True)
                    if "Global" in wb.sheetnames:
                        ws = wb["Global"]
                        metrics = {}
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                            k = (str(row[0].value) if row[0].value is not None else "").strip()
                            v = row[1].value if len(row) > 1 else None
                            if k:
                                metrics[k] = v
                        def _get(*keys):
                            low = {kk.lower().replace(" ", ""): kk for kk in metrics}
                            for k in keys:
                                if k in metrics and metrics[k] is not None:
                                    return metrics[k]
                                t = k.lower().replace(" ", "")
                                if t in low and metrics.get(low[t]) is not None:
                                    return metrics[low[t]]
                            return None
                        def _as_float(x):
                            if x is None: return None
                            s = str(x).strip().replace(",", ".").replace("%", "")
                            try: return float(s)
                            except: return None
                        # Total trades (entier), winrate global (0..1), TP2 si dispo
                        tr = _as_float(_get("Total Trades","Total Trad"))
                        wrg = _as_float(_get("Winrate Global","Winrate Gl","WinrateGL","WinrateGlobal","Winrate TP1"))
                        wr2 = _as_float(_get("TP2 Winrate","Winrate TP2"))
                        trades_count = int(tr) if tr is not None else None
                        wr1 = (wrg/100.0) if wrg is not None else None
                        wr2 = (wr2/100.0) if wr2 is not None else None
                        # ✅ si wr2 toujours None, on tente l’onglet 'TP2_Global'
                        if wr2 is None and "TP2_Global" in wb.sheetnames:
                            ws2 = wb["TP2_Global"]
                            metrics2 = {}
                            for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
                                k = (str(row[0].value) if row[0].value is not None else "").strip()
                                v = row[1].value if len(row) > 1 else None
                                if k:
                                    metrics2[k] = v
                            def _get2(*keys):
                                low = {kk.lower().replace(" ", ""): kk for kk in metrics2}
                                for k in keys:
                                    if k in metrics2 and metrics2[k] is not None:
                                        return metrics2[k]
                                    t = k.lower().replace(" ", "")
                                    if t in low and metrics2.get(low[t]) is not None:
                                        return metrics2[low[t]]
                                return None
                            wr2b = _as_float(_get2("TP2 Winrate","Winrate TP2"))
                            if wr2b is not None:
                                wr2 = wr2b/100.0
                    wb.close()
            except Exception:
                pass

        # label construit avec le pair/symbol corrigé
        params_for_label = dict(params)
        if pair and not params_for_label.get("pair"): params_for_label["pair"] = pair
        if symbol and not params_for_label.get("symbol"): params_for_label["symbol"] = symbol
        

        label, period = _compose_label(params_for_label, run_dir)

        # ✅ si pas trouvé dans params ni _compose_label, on tente dans le nom de dossier
        if not period:
            period = _extract_period_from_dirname(run_dir.name) or ""

        item = CompareOptionsItem(
            id=run_dir.name,
            label=label,
            pair=pair,
            symbol=symbol,
            timeframe=(params.get("timeframe") or "").upper(),
            period=period or params.get("period") or "",
            strategy=params.get("strategy") or "",
            created_at=None,
            trades_count=trades_count,
            winrate_tp1=wr1,
            winrate_tp2=wr2,
            # alignement dashboard
            winrate=params.get("winrate"),
            trades=params.get("trades"),
            metrics=params.get("metrics"),
            xlsx_filename=params.get("xlsx_filename"),
            folder=run_dir.name,
        )
        items.append(item)

    return CompareOptionsResponse(items=items)

# ------------ API: séries pour graph ------------

def build_compare_series(current_user_id: str, req: CompareDataRequest) -> CompareDataResponse:
    metric = req.metric

    # Déterminer buckets & type de valeur
    if metric == "session":
        buckets = DEFAULT_SESSIONS; value_type = "percentage"; precision = 1
    elif metric == "day":
        buckets = DEFAULT_DAYS;     value_type = "percentage"; precision = 1
    elif metric == "hour":
        buckets = DEFAULT_HOURS;    value_type = "percentage"; precision = 1
    elif metric in {"winrate_tp1", "winrate_tp2", "sl_rate"}:
        buckets = ["Global"];       value_type = "percentage"; precision = 2
    elif metric == "trades_count":
        buckets = ["Global"];       value_type = "count";      precision = 0
    else:
        buckets = ["Global"];       value_type = "percentage"; precision = 2

    series: List[SeriesItem] = []

    # Map id -> dossier (perfs)
    id_to_dir: Dict[str, Path] = {d.name: d for d in _find_runs(ANALYSIS_DIR)}

    for analysis_id in req.analysis_ids:
        run_dir = id_to_dir.get(analysis_id)
        if not run_dir:
            series.append(SeriesItem(analysis_id=analysis_id, label=f"{analysis_id}", values=[None]*len(buckets)))
            continue

        params = _load_params(run_dir)
        if not _own_by_user(params, current_user_id):
            series.append(SeriesItem(analysis_id=analysis_id, label=f"{analysis_id}", values=[None]*len(buckets)))
            continue

        files = _detect_files(run_dir)
        # ✅ même label que la liste: normalise pair/symbol puis compose
        pair, symbol = _normalize_pair_symbol(params, run_dir)
        params_for_label = dict(params)
        if pair and not params_for_label.get("pair"): params_for_label["pair"] = pair
        if symbol and not params_for_label.get("symbol"): params_for_label["symbol"] = symbol
        label, _period = _compose_label(params_for_label, run_dir)
 
        values: List[Optional[float]] = [None]*len(buckets)

        if metric == "session" and files["sessions"]:
            df = _safe_read_csv(files["sessions"])
            if df is not None and {"session", "winrate"}.issubset(df.columns):
                wr_map = {str(r["session"]): float(r["winrate"])/100.0 for _, r in df.iterrows()}
                values = [wr_map.get(b, None) for b in buckets]

        elif metric == "day" and files["day"]:
            df = _safe_read_csv(files["day"])
            if df is not None:
                # colonnes tolérées
                # label du jour
                day_col_candidates = ["day_name", "day", "jour", "weekday", "jour_name"]
                day_col = next((c for c in day_col_candidates if c in df.columns), None)
                # winrate
                wr_col_candidates  = ["winrate", "winrate_tp1", "winrate (%)", "wr", "winrate_global"]
                wr_col = next((c for c in wr_col_candidates if c in df.columns), None)
                if day_col and wr_col:
                    wr_map = {}
                    for _, r in df.iterrows():
                        code = _normalize_day_label(str(r[day_col]))
                        try:
                            val = float(str(r[wr_col]).replace("%", "").strip())/100.0
                        except Exception:
                            val = None
                        if code:
                            wr_map[code] = val
                    values = [wr_map.get(b, None) for b in buckets]

        elif metric == "hour" and files["hour"]:
            df = _safe_read_csv(files["hour"])
            if df is not None and {"hour", "winrate"}.issubset(df.columns):
                wr_map = {f"{int(r['hour']):02d}": float(r["winrate"])/100.0 for _, r in df.iterrows()}
                values = [wr_map.get(b, None) for b in buckets]

        elif metric in {"winrate_tp1", "winrate_tp2", "sl_rate", "trades_count"}:
            got = False
            # 1) CSV global si présent
            if files["global"]:
                df = _safe_read_csv(files["global"])
                if df is not None and {"Metric", "Value"}.issubset(df.columns):
                    # dictionnaire normalisé -> valeur
                    kv = {}
                    for _, r in df.iterrows():
                        k = _norm_key(r.get("Metric", ""))
                        v = r.get("Value", None)
                        if v is None: continue
                        try:
                            v = float(str(v).replace("%", "").strip())
                        except Exception:
                            continue
                        if k: kv[k] = v
                    def pick(*cands: str) -> Optional[float]:
                        for c in cands:
                            v = kv.get(_norm_key(c))
                            if v is not None:
                                return v
                        return None

                    if metric == "trades_count":
                        total_trades = pick("Total Trades", "Trades", "Nombre Trades", "Total")
                        values = [float(total_trades) if total_trades is not None else None]; got = True

                    elif metric == "winrate_tp1":
                        wr = pick("Winrate Global", "Winrate TP1", "TP1 Winrate", "Winrate")
                        values = [wr/100.0 if wr is not None else None]; got = True

                    elif metric == "winrate_tp2":
                        # 1) Essai direct via libellés tolérants
                        wr = pick("TP2 Winrate", "Winrate TP2", "TP2 (%)", "TP2 Rate", "WR TP2")
                        if wr is not None:
                            values = [wr/100.0]; got = True
                        else:
                            # 1-bis) Si un CSV dédié TP2 existe, on le lit avant tout fallback calculé
                            if not got and files.get("tp2_global"):
                                df_tp2 = _safe_read_csv(files["tp2_global"])
                                if df_tp2 is not None and {"Metric", "Value"}.issubset(df_tp2.columns):
                                    try:
                                        row = df_tp2.loc[df_tp2["Metric"].astype(str).str.strip().str.lower().replace({"winrate tp2":"winrate tp2"}) == "winrate tp2", "Value"]
                                    except Exception:
                                        row = df_tp2.loc[df_tp2["Metric"].astype(str) == "Winrate TP2", "Value"]
                                    if not row.empty:
                                        try:
                                            v = float(str(row.iloc[0]).replace("%", "").strip())/100.0
                                            values = [v]; got = True
                                        except Exception:
                                            pass
                            # 2) Fallback calculé (dernier recours) si toujours rien
                            if not got:
                                total = pick("Total Trades", "Trades", "Nombre Trades", "Total")
                                tp1   = pick("TP1") or 0.0
                                tp2   = pick("TP2") or 0.0
                                sl    = pick("SL")  or 0.0
                                if total is None or total <= 0:
                                    total = tp1 + tp2 + sl
                                rate = (tp2/total) if total and total > 0 else None
                                values = [rate]; got = True
 
                    elif metric == "sl_rate":
                        # ✅ calcule sur Total Trades (fallback TP1+TP2+SL)
                        total = pick("Total Trades", "Trades", "Nombre Trades", "Total")
                        tp1   = pick("TP1") or 0.0
                        tp2   = pick("TP2") or 0.0
                        sl    = pick("SL")  or 0.0
                        if total is None or total <= 0:
                            total = tp1 + tp2 + sl
                        rate = (sl/total) if total and total > 0 else None
                        values = [rate]; got = True
          
            # 2) Fallback XLSX (onglet "Global") si pas de CSV ou valeurs manquantes
            if not got:
                try:
                    candidates = list(run_dir.glob("analyse_*_resultats.xlsx"))
                    if candidates:
                        wb = openpyxl.load_workbook(candidates[0], data_only=True)
                        if "Global" in wb.sheetnames:
                            ws = wb["Global"]
                            metrics = {}
                            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                                k = (str(row[0].value) if row[0].value is not None else "").strip()
                                v = row[1].value if len(row) > 1 else None
                                if k:
                                    metrics[k] = v
                            def _get(*keys):
                                low = {kk.lower().replace(" ", ""): kk for kk in metrics}
                                for k in keys:
                                    if k in metrics and metrics[k] is not None:
                                        return metrics[k]
                                    t = k.lower().replace(" ", "")
                                    if t in low and metrics.get(low[t]) is not None:
                                        return metrics[low[t]]
                                return None
                            def _f(x):
                                if x is None: return None
                                s = str(x).strip().replace(",", ".").replace("%", "")
                                try: return float(s)
                                except: return None
                            if metric == "trades_count":
                                tr = _f(_get("Total Trades","Total Trad"))
                                values = [float(tr) if tr is not None else None]
                            elif metric == "winrate_tp1":
                                wr = _f(_get("Winrate Global","Winrate TP1","TP1 Winrate","Winrate","WinrateTP1","WinrateGlobal"))
                                values = [wr/100.0 if wr is not None else None]
                            elif metric == "winrate_tp2":
                                # 1) Essai via libellés tolérants
                                wr = _f(_get("TP2 Winrate","Winrate TP2","TP2 (%)","TP2 Rate","WR TP2"))
                                if wr is not None:
                                    values = [wr/100.0]
                                else:
                                    # 1-bis) Essai via feuille 'TP2_Global' si présente
                                    if "TP2_Global" in wb.sheetnames:
                                        ws2 = wb["TP2_Global"]
                                        metrics2 = {}
                                        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
                                            k = (str(row[0].value) if row[0].value is not None else "").strip()
                                            v = row[1].value if len(row) > 1 else None
                                            if k:
                                                metrics2[k] = v
                                        def _get2(*keys):
                                            low = {kk.lower().replace(" ", ""): kk for kk in metrics2}
                                            for k in keys:
                                                if k in metrics2 and metrics2[k] is not None:
                                                    return metrics2[k]
                                                t = k.lower().replace(" ", "")
                                                if t in low and metrics2.get(low[t]) is not None:
                                                    return metrics2[low[t]]
                                            return None
                                        wr_bis = _f(_get2("TP2 Winrate","Winrate TP2","TP2 (%)","TP2 Rate","WR TP2"))
                                        if wr_bis is not None:
                                            values = [wr_bis/100.0]
                                            got = True
                                    if not got:

                                        # 2) Fallback calculé sur les comptages
                                        total = _f(_get("Total Trades","Trades","Nombre Trades","Total"))
                                        tp1   = _f(_get("TP1")) or 0.0
                                        tp2   = _f(_get("TP2")) or 0.0
                                        sl    = _f(_get("SL"))  or 0.0
                                        if total is None or total <= 0:
                                            total = (tp1 + tp2 + sl)
                                        rate = (tp2/total) if total and total > 0 else None
                                        values = [rate]
                            elif metric == "sl_rate":
                                total = _f(_get("Total Trades","Trades","Nombre Trades","Total"))
                                tp1   = _f(_get("TP1")) or 0.0
                                tp2   = _f(_get("TP2")) or 0.0
                                sl    = _f(_get("SL"))  or 0.0
                                if total is None or total <= 0:
                                    total = (tp1 + tp2 + sl)
                                rate = (sl/total) if total and total > 0 else None
                                values = [rate]
                        wb.close()
                except Exception:
                    pass

        series.append(SeriesItem(analysis_id=analysis_id, label=label, values=values))

    return CompareDataResponse(
        metric=metric,
        value_type=value_type,
        precision=precision,
        buckets=buckets,
        series=series
    )
