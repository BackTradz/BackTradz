"""
File: backend/routes/user_dashboard_routes.py
Role: Récupère les backtests d'un utilisateur pour affichage dans son dashboard.
Depends:
  - backend.auth.get_current_user (auth X-API-Key)
  - backend/data/analysis/<...>/*.json (métadonnées de run)
  - openpyxl (lecture 'Global' pour winrate)
Side-effects:
  - Lecture disque (JSON + XLSX)
Security:
  - Protégé par get_current_user (header X-API-Key)
Notes:
  - Les prints '✅/📄/⛔️/📊' sont des logs debug utiles, conservés.
"""

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from pathlib import Path
from backend.core.paths import ANALYSIS_DIR, DATA_ROOT
from backend.auth import get_current_user, get_user_by_token  # get_user_by_token si besoin
import json
from typing import List
import openpyxl
import os
from fastapi.responses import JSONResponse
import shutil




router = APIRouter()

@router.get("/user/backtests")
def get_user_backtests(request: Request, user=Depends(get_current_user)):
    """
    Scanne backend/data/analysis/** pour les runs de l'utilisateur connecté
    et renvoie des items prêts à afficher (incluant metrics détaillées).
    """
    print(f"✅ USER CONNECTÉ → {user.id}")
    base_dir = ANALYSIS_DIR
    alt_dir = Path("backend/data/analysis").resolve()

    search_dirs = []
    if base_dir.exists():
        search_dirs.append(base_dir)
    if alt_dir.exists() and alt_dir != base_dir:
        search_dirs.append(alt_dir)

    backtests = []

    # Parcours récursif de tous les JSON d'analyse dans TOUTES les racines
    for root in search_dirs:
        for subdir in root.glob("**/*.json"):
            try:
                with open(subdir, "r", encoding="utf-8") as f:
                    data = json.load(f)

                # dédupe par dossier d'analyse (miroir /var vs backend/)
                folder_name = subdir.parent.name
                if folder_name in seen_folders:
                    continue
                seen_folders.add(folder_name)

                print(f"📄 SCANNING {subdir.name} | user_id in file: {data.get('user_id')}")
                if data.get("user_id") != user.id:
                    #print("⛔️ USER MISMATCH → ignoré")
                    continue

                #print("✅ USER MATCH → ajouté au dashboard")

                # --------- Fallbacks champs manquants depuis le nom de dossier ----------
                symbol   = data.get("pair")
                timeframe = data.get("timeframe")
                period   = data.get("period") or ""
                strategy = data.get("strategy") or ""
                sl_pips  = data.get("params", {}).get("sl_pips", 100)

                folder_name = subdir.parent.name  # ex: AUDUSD_M5_ob_pullback_pure_2025-07-01to2025-07-31_sl100
                try:
                    parts = folder_name.split("_")
                    if not symbol and len(parts) >= 1:
                        symbol = parts[0]
                    if not timeframe and len(parts) >= 2 and parts[1].upper() in {"M1","M5","M15","M30","H1","H4","D1"}:
                        timeframe = parts[1].upper()
                    # 🔥 fallback pour la période
                    if not period:
                        for p in parts:
                            if "to" in p and any(ch.isdigit() for ch in p):
                                period = p.replace("__", "_")
                                break
                except Exception:
                    pass
                        

                # ---------------- Résolution tolérante du fichier XLSX ------------------
                filename = f"analyse_{strategy}_{symbol}_SL{sl_pips}_{period}_resultats.xlsx"
                xlsx_path = subdir.parent / filename

                if not xlsx_path.exists():
                    variants = set()
                    if period:
                        variants.update({
                            period,
                            period.replace("_to_", " to "),
                            period.replace(" to ", "_to_"),
                            period.replace(" ", "_"),
                            period.replace("_", " "),
                            period.replace("to", " to ").replace("__", "_"),
                        })

                    # essaie les variantes de period
                    for per in variants:
                        cand = subdir.parent / f"analyse_{strategy}_{symbol}_SL{sl_pips}_{per}_resultats.xlsx"
                        if cand.exists():
                            xlsx_path, filename = cand, cand.name
                            break

                    # fallback: prends le 1er résultat plausible
                    if not xlsx_path.exists():
                        cands = list(subdir.parent.glob(f"analyse_{strategy}_{symbol}_SL*_*resultats.xlsx")) \
                                or list(subdir.parent.glob("analyse_*_resultats.xlsx"))
                        if cands:
                            xlsx_path, filename = cands[0], cands[0].name

                # ---------------------- Lecture de la feuille Global ---------------------
                winrate = "N/A"
                trades  = None
                metrics_payload = None

                if xlsx_path and xlsx_path.exists():
                    try:
                        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                        if "Global" in wb.sheetnames:
                            ws = wb["Global"]

                            # indexe toutes les métriques
                            metrics = {}
                            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                                key = (str(row[0].value) if row[0].value is not None else "").strip()
                                val = row[1].value if len(row) > 1 else None
                                if key:
                                    metrics[key] = val

                            # helpers
                            def get_metric(*keys):
                                for k in keys:
                                    if k in metrics and metrics[k] is not None:
                                        return metrics[k]
                                low = {k.lower().replace(" ", ""): k for k in metrics.keys()}
                                for k in keys:
                                    t = k.lower().replace(" ", "")
                                    if t in low:
                                        return metrics[low[t]]
                                return None

                            def fmt_pct(x):
                                if x is None: return None
                                s = str(x).strip().replace(",", ".")
                                try:
                                    n = float(s)
                                    return f"{n*100:.2f}%" if n <= 1 else f"{n:.2f}%"
                                except:
                                    return s if s.endswith("%") else f"{s}%"

                            def fmt_num(x):
                                if x is None: return None
                                s = str(x).strip().replace(",", ".")
                                try:
                                    return round(float(s), 2)
                                except:
                                    return None

                            def fmt_int(x):
                                if x is None: return None
                                s = str(x).strip().replace(",", ".")
                                try:
                                    return int(float(s))
                                except:
                                    return None

                            # winrate (priorité Global > TP1 > générique hors buy/sell)
                            wr = get_metric("Winrate Gl", "Winrate Global", "WinrateGL", "WinrateGlobal")
                            if wr is None:
                                wr = get_metric("Winrate TP1", "TP1 Winrate")
                            if wr is None:
                                for k, v in metrics.items():
                                    kl = k.lower()
                                    if "winrate" in kl and "buy" not in kl and "sell" not in kl and v is not None:
                                        wr = v
                                        break
                            if wr is not None:
                                winrate = fmt_pct(wr)

                            # trades
                            tr = get_metric("Total Trades", "Total Trad")
                            if tr is not None:
                                trades = fmt_int(tr)

                            # payload détaillé pour le modal
                            metrics_payload = {
                                "winrate_global": fmt_pct(get_metric("Winrate Gl","Winrate Global","WinrateGL","WinrateGlobal")),
                                "buy_winrate":  fmt_pct(get_metric("Buy Winrate","BuyWinrate","Winrate Buy","WinrateBuy")),
                                "sell_winrate": fmt_pct(get_metric("Sell Winrate","SellWinrate","Winrate Sell","WinrateSell")),
                                "pct_buy":        fmt_pct(get_metric("% Buy","%Buy")),
                                "pct_sell":       fmt_pct(get_metric("% Sell","%Sell")),
                                "tp1":            fmt_int(get_metric("TP1")),
                                "sl":             fmt_int(get_metric("SL")),
                                "rr_tp1":         fmt_num(get_metric("RR TP1 (avg)","RR TP1")),
                                "rr_tp2":         fmt_num(get_metric("RR TP2 (avg)","RR TP2")),
                            "sl_size":  fmt_num(get_metric("SL Size (avg,pips)", "SL Size", "Avg SL size", "SL (avg size)", "SL size avg")),
                                "tp1_size": fmt_num(get_metric("TP1 Size (avg,pips)", "TP1 Size", "Avg TP1 size", "TP1 (avg size)", "TP1 size avg")),
                            }
                        wb.close()
                    except Exception as e:
                        print(f"❌ Erreur lecture xlsx {xlsx_path}: {e}")

                # -------------------------- Ajout de l'item -----------------------------
                backtests.append({
                    "strategy": strategy,
                    "symbol": symbol,
                    "timeframe": timeframe,
                    "period": period,
                    "winrate": winrate,
                    "trades": trades,
                    "xlsx_filename": filename,
                    "folder": str(subdir.parent.name),
                    "metrics": metrics_payload,   # <— détails pour le modal
                })

            except Exception as e:
                print(f"❌ Erreur lecture {subdir.name} → {e}")
                continue

    print(f"📊 TOTAL BACKTESTS TROUVÉS : {len(backtests)}")
    return backtests




@router.delete("/user/backtests/delete/{folder}")
def delete_backtest(folder: str, user=Depends(get_current_user)):
    """
    Supprime un dossier d'analyse uniquement si l'utilisateur est propriétaire.

    Args:
      folder (str): nom du dossier à supprimer (UUID ou nom généré)

    Returns:
      dict: message de confirmation
    """

    target_path = (ANALYSIS_DIR / folder).resolve()

    if not target_path.exists() or not target_path.is_dir():
        raise HTTPException(status_code=404, detail="Dossier introuvable")

    json_files = list(target_path.glob("*.json"))
    if not json_files:
        raise HTTPException(status_code=400, detail="Aucun fichier JSON trouvé dans le dossier")

    try:
        with open(json_files[0], "r") as f:
            data = json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture JSON: {e}")

    if data.get("user_id") != user.id:
        raise HTTPException(status_code=403, detail="Non autorisé à supprimer ce backtest")

    try:
        shutil.rmtree(target_path)
        print(f"🗑️ Dossier supprimé : {folder}")
        return JSONResponse(content={"message": "Backtest supprimé avec succès"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur suppression dossier: {e}")

# backend/routes/user_dashboard_routes.py


# === CSV deletion endpoints (tolérants) ===============================
from fastapi import Query, Body
from fastapi.responses import Response

def _normalize_backend_rel(p: str) -> str:
    x = (p or "").replace("\\", "/").strip()
    return x


def _delete_csv_file(rel_after_backend: str, user):
    """Supprime le fichier situé sous backend/<rel_after_backend>."""
    # Autorise uniquement des fichiers sous DATA_ROOT (output / output_live / analysis)
    target = Path(rel_after_backend)
    if not target.is_absolute():
        target = (DATA_ROOT / rel_after_backend).resolve()
    if DATA_ROOT not in target.parents and target != DATA_ROOT:
        raise HTTPException(status_code=400, detail="Chemin invalide")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="CSV introuvable")

    try:
        target.unlink()
        print(f"🗑️ CSV supprimé : {target}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur suppression: {e}")

@router.delete("/delete_csv_by_path", status_code=204)
def delete_csv_by_query(path: str = Query(..., description="Chemin incluant 'backend/'"),
                        user=Depends(get_current_user)):
    rel = _normalize_backend_rel(path)
    _delete_csv_file(rel, user)
    return Response(status_code=204)

@router.post("/delete_csv_by_path", status_code=204)
def delete_csv_by_post(payload: dict = Body(...), user=Depends(get_current_user)):
    rel = _normalize_backend_rel(payload.get("path", ""))
    if not rel:
        raise HTTPException(status_code=400, detail="Paramètre 'path' manquant")
    _delete_csv_file(rel, user)
    return Response(status_code=204)

@router.delete("/delete_csv_by_path/{path:path}", status_code=204)
def delete_csv_by_pathparam(path: str, user=Depends(get_current_user)):
    rel = _normalize_backend_rel(path)
    _delete_csv_file(rel, user)
    return Response(status_code=204)

